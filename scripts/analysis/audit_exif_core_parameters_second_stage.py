"""Second-stage audit of four core EXIF parameters for the NYHA project.

This script is intentionally read-only with respect to source images, labels,
splits, historical experiments, and training code.  It writes a new audit
package under reports/exif_core_parameter_second_stage_audit.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import platform
import re
import subprocess
import sys
import warnings
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scipy
import PIL
import openpyxl
from PIL import Image
from matplotlib.colors import ListedColormap
from openpyxl import load_workbook
from scipy import ndimage, stats
from scipy.cluster import hierarchy
from scipy.spatial.distance import squareform
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_selection import mutual_info_classif
from sklearn.linear_model import LinearRegression, LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    balanced_accuracy_score,
    confusion_matrix,
    f1_score,
    precision_score,
    recall_score,
    roc_auc_score,
    roc_curve,
)
from sklearn.preprocessing import OneHotEncoder, StandardScaler
import sklearn
import statsmodels.api as sm
import statsmodels


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SEED = 2026
BOOTSTRAP_REPEATS = 2000
CLASS_NAMES = {0: "normal", 1: "mild", 2: "severe"}
SEX_NAMES = {0: "female", 1: "male"}
CORE_FIELDS = ["ExposureTime", "FNumber", "ISOSpeedRatings", "BrightnessValue"]
AUX_FIELDS = ["Make", "Model", "MeteringMode", "Flash", "ShutterSpeedValue"]
DERIVED_FIELDS = [
    "log2_exposure_time",
    "log2_iso_gain",
    "log2_fnumber",
    "aperture_value_from_f",
    "time_value_from_t",
    "EV100",
    "relative_optical_exposure",
    "combined_exposure_gain",
    "sensitivity_value",
    "brightness_apex_pred",
    "brightness_residual",
    "device_centered_brightness",
]
ANALYSIS_FIELDS = CORE_FIELDS + DERIVED_FIELDS
DEVICE_MODEL_DERIVED = [field for field in DERIVED_FIELDS if field != "device_centered_brightness"]
POSITIVE_LOG_PAIRS = {
    "ExposureTime": "log2_exposure_time",
    "FNumber": "log2_fnumber",
    "ISOSpeedRatings": "log2_iso_gain",
}
PALETTE = {
    "HONOR/BVL-AN00": "#3B6FB6",
    "Xiaomi/M2006J10C": "#D9863D",
    "normal": "#5B8FF9",
    "mild": "#61B15A",
    "severe": "#D95F59",
}


def configure_plotting() -> None:
    matplotlib.rcParams.update(
        {
            "font.family": "sans-serif",
            "font.sans-serif": ["Arial", "Helvetica", "DejaVu Sans", "sans-serif"],
            "svg.fonttype": "none",
            "pdf.fonttype": 42,
            "font.size": 8,
            "axes.spines.right": False,
            "axes.spines.top": False,
            "axes.linewidth": 0.8,
            "legend.frameon": False,
            "figure.facecolor": "white",
            "axes.facecolor": "white",
        }
    )


def save_figure(fig: plt.Figure, base_path: Path, dpi: int = 300) -> None:
    fig.savefig(base_path.with_suffix(".png"), dpi=dpi, bbox_inches="tight", facecolor="white")
    fig.savefig(base_path.with_suffix(".svg"), bbox_inches="tight", facecolor="white")
    plt.close(fig)


def save_csv(frame: pd.DataFrame, path: Path) -> None:
    frame.to_csv(path, index=False, encoding="utf-8-sig")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.startswith("'"):
        text = text[1:].strip()
    return text


def parse_number(value: Any) -> float:
    text = clean_text(value).replace(",", "")
    if not text or text.casefold() in {"none", "nan", "na", "n/a", "null", "unknown"}:
        return float("nan")
    try:
        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?/[-+]?\d+(?:\.\d+)?", text):
            numerator, denominator = text.split("/", 1)
            denominator_value = float(denominator)
            return float(numerator) / denominator_value if denominator_value else float("nan")
        value_float = float(text)
        return value_float if math.isfinite(value_float) else float("nan")
    except (TypeError, ValueError):
        return float("nan")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def sha256_directory(path: Path, pattern: str = "*.png") -> dict[str, Any]:
    files = sorted(path.glob(pattern), key=lambda item: item.name.casefold())
    manifest_digest = hashlib.sha256()
    total_bytes = 0
    for file_path in files:
        file_digest = sha256_file(file_path)
        size = file_path.stat().st_size
        total_bytes += size
        manifest_digest.update(f"{file_path.name}\t{size}\t{file_digest}\n".encode("utf-8"))
    return {
        "path": str(path.resolve()),
        "pattern": pattern,
        "file_count": len(files),
        "total_bytes": total_bytes,
        "content_manifest_sha256": manifest_digest.hexdigest(),
    }


def markdown_table(frame: pd.DataFrame, max_rows: int | None = None) -> str:
    view = frame.copy()
    if max_rows is not None:
        view = view.head(max_rows)
    if view.empty:
        return "（无记录）"
    view = view.fillna("")
    columns = [str(column) for column in view.columns]
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join(["---"] * len(columns)) + " |",
    ]
    for row in view.itertuples(index=False, name=None):
        values = []
        for value in row:
            if isinstance(value, float):
                rendered = f"{value:.6g}" if math.isfinite(value) else ""
            else:
                rendered = str(value)
            values.append(rendered.replace("|", "\\|"))
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def load_exif_values(workbook_path: Path) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    target = set(CORE_FIELDS + AUX_FIELDS)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    if len(sheet_names) < 3:
        raise ValueError(f"Metadata workbook has only {len(sheet_names)} sheets: {sheet_names}")

    metadata_sheet = workbook.worksheets[1]
    iterator = metadata_sheet.iter_rows(values_only=True)
    headers = [clean_text(value) or f"unnamed_{index}" for index, value in enumerate(next(iterator))]
    metadata = pd.DataFrame(list(iterator), columns=headers)
    if "ID" not in metadata.columns:
        raise ValueError(f"ID missing from metadata sheet; columns={metadata.columns.tolist()}")
    metadata["ID"] = metadata["ID"].map(clean_text)
    metadata = metadata.loc[metadata["ID"] != ""].copy()

    raw_sheet = workbook.worksheets[2]
    iterator = raw_sheet.iter_rows(values_only=True)
    next(iterator)
    values: dict[str, dict[str, Any]] = defaultdict(dict)
    observed: dict[tuple[str, str], list[str]] = defaultdict(list)
    for row in iterator:
        if len(row) < 6:
            continue
        image_id = clean_text(row[0])
        parameter = clean_text(row[4])
        if not image_id or parameter not in target:
            continue
        raw_value = row[5]
        observed[(image_id, parameter)].append(clean_text(raw_value))
        if parameter not in values[image_id] or not clean_text(values[image_id][parameter]):
            values[image_id][parameter] = raw_value
    workbook.close()

    conflict_rows = []
    for (image_id, parameter), entries in observed.items():
        unique_values = sorted(set(entries))
        if len(unique_values) > 1:
            conflict_rows.append(
                {"ID": image_id, "parameter": parameter, "values": " | ".join(unique_values)}
            )

    records = []
    for image_id in metadata["ID"].astype(str):
        item: dict[str, Any] = {"ID": image_id}
        for field in CORE_FIELDS + AUX_FIELDS:
            raw_value = values.get(image_id, {}).get(field)
            if field in {"Make", "Model"}:
                item[field] = clean_text(raw_value)
            else:
                item[field] = parse_number(raw_value)
                item[f"{field}_raw"] = clean_text(raw_value)
        records.append(item)
    frame = pd.DataFrame(records)
    return frame, conflict_rows


def read_split(path: Path, cohort_name: str) -> pd.DataFrame:
    frame = pd.read_csv(
        path,
        dtype={"ID": "string", "patient_group_id": "string"},
        encoding="utf-8-sig",
    )
    required = {"ID", "patient_group_id", "SEX", "NYHA", "label_3class", "fold"}
    missing = required.difference(frame.columns)
    if missing:
        raise ValueError(f"{cohort_name} split is missing columns: {sorted(missing)}")
    frame = frame.copy()
    frame["ID"] = frame["ID"].astype(str).str.strip()
    frame["patient_group_id"] = frame["patient_group_id"].astype(str).str.strip()
    for field in ["SEX", "NYHA", "label_3class", "fold"]:
        frame[field] = pd.to_numeric(frame[field], errors="raise").astype(int)
    frame["label_3class_name"] = frame["label_3class"].map(CLASS_NAMES)
    frame["sex_name"] = frame["SEX"].map(SEX_NAMES)
    frame["cohort"] = cohort_name
    return frame


def derive_parameters(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    exposure = result["ExposureTime"].where(result["ExposureTime"] > 0)
    fnumber = result["FNumber"].where(result["FNumber"] > 0)
    iso = result["ISOSpeedRatings"].where(result["ISOSpeedRatings"] > 0)
    result["log2_exposure_time"] = np.log2(exposure)
    result["log2_iso_gain"] = np.log2(iso / 100.0)
    result["log2_fnumber"] = np.log2(fnumber)
    result["aperture_value_from_f"] = 2.0 * result["log2_fnumber"]
    result["time_value_from_t"] = -result["log2_exposure_time"]
    result["EV100"] = result["aperture_value_from_f"] + result["time_value_from_t"]
    result["relative_optical_exposure"] = (
        result["log2_exposure_time"] - 2.0 * result["log2_fnumber"]
    )
    result["combined_exposure_gain"] = (
        result["relative_optical_exposure"] + result["log2_iso_gain"]
    )
    result["sensitivity_value"] = np.log2(iso / 3.125)
    result["brightness_apex_pred"] = (
        result["aperture_value_from_f"]
        + result["time_value_from_t"]
        - result["sensitivity_value"]
    )
    result["brightness_residual"] = (
        result["BrightnessValue"] - result["brightness_apex_pred"]
    )
    result["shutter_time_apex_error"] = (
        result["ShutterSpeedValue"] - result["time_value_from_t"]
    )
    centered = pd.Series(index=result.index, dtype=float)
    for camera_id, indices in result.groupby("camera_id").groups.items():
        values = result.loc[indices, "BrightnessValue"].astype(float)
        median = values.median()
        iqr = values.quantile(0.75) - values.quantile(0.25)
        scale = iqr if math.isfinite(iqr) and iqr > 0 else 1.0
        centered.loc[indices] = (values - median) / scale
    result["device_centered_brightness"] = centered
    return result


def median_absolute_deviation(values: pd.Series | np.ndarray) -> float:
    array = np.asarray(values, dtype=float)
    array = array[np.isfinite(array)]
    if len(array) == 0:
        return float("nan")
    median = float(np.median(array))
    return float(np.median(np.abs(array - median)))


def describe_values(values: pd.Series, total_n: int | None = None) -> dict[str, Any]:
    numeric = pd.to_numeric(values, errors="coerce")
    valid = numeric[np.isfinite(numeric.to_numpy(dtype=float))]
    if total_n is None:
        total_n = len(numeric)
    result: dict[str, Any] = {
        "total_n": int(total_n),
        "n": int(len(valid)),
        "missing_or_invalid_n": int(total_n - len(valid)),
        "unique_n": int(valid.nunique()),
    }
    for key in ["min", "p1", "p5", "p25", "median", "p75", "p95", "p99", "max", "mean", "std", "iqr", "mad", "mode", "mode_proportion"]:
        result[key] = float("nan")
    if valid.empty:
        return result
    quantiles = valid.quantile([0.01, 0.05, 0.25, 0.5, 0.75, 0.95, 0.99])
    modes = valid.mode(dropna=True)
    mode = float(modes.iloc[0]) if not modes.empty else float("nan")
    result.update(
        {
            "min": float(valid.min()),
            "p1": float(quantiles.loc[0.01]),
            "p5": float(quantiles.loc[0.05]),
            "p25": float(quantiles.loc[0.25]),
            "median": float(quantiles.loc[0.5]),
            "p75": float(quantiles.loc[0.75]),
            "p95": float(quantiles.loc[0.95]),
            "p99": float(quantiles.loc[0.99]),
            "max": float(valid.max()),
            "mean": float(valid.mean()),
            "std": float(valid.std(ddof=1)) if len(valid) > 1 else 0.0,
            "iqr": float(quantiles.loc[0.75] - quantiles.loc[0.25]),
            "mad": median_absolute_deviation(valid),
            "mode": mode,
            "mode_proportion": float((valid == mode).mean()) if math.isfinite(mode) else float("nan"),
        }
    )
    return result


def patient_group_frame(frame: pd.DataFrame, value_fields: Sequence[str]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for group_id, group in frame.groupby("patient_group_id", sort=True):
        row: dict[str, Any] = {
            "patient_group_id": str(group_id),
            "image_n": int(len(group)),
            "image_ids": ";".join(sorted(group["ID"].astype(str))),
        }
        for field in value_fields:
            row[field] = pd.to_numeric(group[field], errors="coerce").median()
        for field in ["camera_id", "label_3class", "label_3class_name", "SEX", "sex_name", "fold"]:
            values = group[field].dropna().unique().tolist()
            if len(values) == 1:
                row[field] = values[0]
            elif field in {"label_3class", "SEX", "fold"}:
                row[field] = float("nan")
            else:
                row[field] = "AMBIGUOUS"
            row[f"{field}_unique_n"] = len(values)
        rows.append(row)
    return pd.DataFrame(rows)


def summarize_overall(cohorts: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for cohort_name, frame in cohorts.items():
        levels = {
            "image": frame,
            "patient_group_median": patient_group_frame(frame, ANALYSIS_FIELDS),
        }
        for level, level_frame in levels.items():
            for variable in ANALYSIS_FIELDS:
                row = {
                    "cohort": cohort_name,
                    "level": level,
                    "variable": variable,
                    "scale": "log2" if variable in POSITIVE_LOG_PAIRS.values() else "raw_or_derived",
                }
                row.update(describe_values(level_frame[variable]))
                if variable in CORE_FIELDS and level == "image":
                    raw = pd.to_numeric(level_frame[variable], errors="coerce")
                    missing = raw.isna()
                    invalid = (~missing) & (~np.isfinite(raw.astype(float)))
                    if variable in {"ExposureTime", "FNumber", "ISOSpeedRatings"}:
                        invalid |= (~missing) & (raw <= 0)
                    row["missing_n"] = int(missing.sum())
                    row["invalid_n"] = int(invalid.sum())
                else:
                    row["missing_n"] = int(level_frame[variable].isna().sum())
                    row["invalid_n"] = 0
                rows.append(row)
    return pd.DataFrame(rows)


def summarize_grouped(
    cohorts: dict[str, pd.DataFrame],
    group_fields: Sequence[str],
    omit_ambiguous_patient_groups: bool = True,
) -> pd.DataFrame:
    rows = []
    for cohort_name, frame in cohorts.items():
        image_frame = frame.copy()
        patient_frame = patient_group_frame(frame, ANALYSIS_FIELDS)
        levels = {"image": image_frame, "patient_group_median": patient_frame}
        for level, level_frame in levels.items():
            work = level_frame.copy()
            if level == "patient_group_median" and omit_ambiguous_patient_groups:
                for group_field in group_fields:
                    unique_col = f"{group_field}_unique_n"
                    if unique_col in work.columns:
                        work = work.loc[work[unique_col] == 1].copy()
                    elif work[group_field].dtype == object:
                        work = work.loc[work[group_field] != "AMBIGUOUS"].copy()
            for group_values, subset in work.groupby(list(group_fields), dropna=False, sort=True):
                if not isinstance(group_values, tuple):
                    group_values = (group_values,)
                metadata = dict(zip(group_fields, group_values))
                for variable in ANALYSIS_FIELDS:
                    row = {
                        "cohort": cohort_name,
                        "level": level,
                        "variable": variable,
                        **metadata,
                    }
                    row.update(describe_values(subset[variable]))
                    rows.append(row)
    return pd.DataFrame(rows)


def robust_z(values: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(values, errors="coerce").astype(float)
    median = numeric.median()
    mad = median_absolute_deviation(numeric)
    if not math.isfinite(mad) or mad <= 0:
        return pd.Series(np.nan, index=values.index, dtype=float)
    return 0.6744897501960817 * (numeric - median) / mad


def build_outlier_table(frame: pd.DataFrame, formal_ids: set[str]) -> pd.DataFrame:
    rows = []
    for (camera_id,), camera_group in frame.groupby(["camera_id"], sort=True):
        for variable in ANALYSIS_FIELDS:
            values = pd.to_numeric(camera_group[variable], errors="coerce").astype(float)
            valid = values[np.isfinite(values)]
            if valid.empty:
                continue
            q1, q3 = valid.quantile([0.25, 0.75])
            iqr = q3 - q1
            lower = q1 - 1.5 * iqr
            upper = q3 + 1.5 * iqr
            z = robust_z(values)
            iqr_flag = (values < lower) | (values > upper)
            mad_flag = z.abs() > 3.5
            union = iqr_flag | mad_flag
            for index in camera_group.index[union.fillna(False)]:
                image = frame.loc[index]
                rows.append(
                    {
                        "ID": image["ID"],
                        "patient_group_id": image["patient_group_id"],
                        "in_formal_modeling_cohort": image["ID"] in formal_ids,
                        "camera_id": camera_id,
                        "NYHA": image["NYHA"],
                        "label_3class": image["label_3class"],
                        "SEX": image["SEX"],
                        "fold_full": image["fold"],
                        "variable": variable,
                        "value": values.loc[index],
                        "iqr_lower": lower,
                        "iqr_upper": upper,
                        "iqr_outlier": bool(iqr_flag.loc[index]),
                        "robust_z": z.loc[index],
                        "mad_outlier_abs_z_gt_3_5": bool(mad_flag.loc[index]) if pd.notna(mad_flag.loc[index]) else False,
                        "known_first_stage_iso_outlier_abs_z_gt_5": bool(
                            variable == "ISOSpeedRatings" and pd.notna(z.loc[index]) and z.loc[index] > 5.0
                        ),
                    }
                )
    return pd.DataFrame(rows)


def check_cohort_alignment(
    exif: pd.DataFrame,
    full_split: pd.DataFrame,
    formal_split: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    exif_ids = set(exif["ID"])
    full_ids = set(full_split["ID"])
    formal_ids = set(formal_split["ID"])
    full_lookup = full_split.set_index("ID")
    formal_lookup = formal_split.set_index("ID")
    rows = []
    for image_id in sorted(exif_ids | full_ids | formal_ids):
        row: dict[str, Any] = {
            "ID": image_id,
            "in_exif": image_id in exif_ids,
            "in_full_522_split": image_id in full_ids,
            "in_formal_500_split": image_id in formal_ids,
        }
        if image_id in full_ids:
            source = full_lookup.loc[image_id]
            for field in ["patient_group_id", "NYHA", "SEX", "label_3class", "fold"]:
                row[f"full_{field}"] = source[field]
        if image_id in formal_ids:
            source = formal_lookup.loc[image_id]
            for field in ["patient_group_id", "NYHA", "SEX", "label_3class", "fold"]:
                row[f"formal_{field}"] = source[field]
            row["formal_metadata_agrees_with_full"] = all(
                str(row.get(f"formal_{field}")) == str(row.get(f"full_{field}"))
                for field in ["patient_group_id", "NYHA", "SEX", "label_3class"]
            )
        else:
            row["formal_metadata_agrees_with_full"] = np.nan
        rows.append(row)
    alignment = pd.DataFrame(rows)

    expected_mapping = {0: 0, 1: 1, 2: 1, 3: 2, 4: 2}
    checks: list[dict[str, Any]] = []

    def add(name: str, passed: bool, detail: str) -> None:
        checks.append({"check": name, "status": "PASS" if passed else "FAIL", "detail": detail})

    add("exif_unique_id", exif["ID"].is_unique, f"rows={len(exif)}, unique={exif['ID'].nunique()}")
    add("full_split_unique_id", full_split["ID"].is_unique, f"rows={len(full_split)}, unique={full_split['ID'].nunique()}")
    add("formal_split_unique_id", formal_split["ID"].is_unique, f"rows={len(formal_split)}, unique={formal_split['ID'].nunique()}")
    add("exif_equals_full_split", exif_ids == full_ids, f"exif_only={len(exif_ids-full_ids)}, split_only={len(full_ids-exif_ids)}")
    add("formal_subset_of_full", formal_ids <= full_ids, f"formal_not_full={len(formal_ids-full_ids)}")
    add("formal_has_exif", formal_ids <= exif_ids, f"formal_without_exif={len(formal_ids-exif_ids)}")
    add(
        "formal_metadata_agreement",
        bool(alignment.loc[alignment["in_formal_500_split"], "formal_metadata_agrees_with_full"].all()),
        "patient_group_id, NYHA, SEX and label_3class compared for all formal IDs",
    )
    for cohort_name, split in [("full", full_split), ("formal", formal_split)]:
        leakage = split.groupby("patient_group_id")["fold"].nunique().gt(1)
        add(
            f"{cohort_name}_patient_group_no_cross_fold",
            not leakage.any(),
            f"leaking_groups={leakage[leakage].index.astype(str).tolist()}",
        )
        mapping_ok = split.apply(
            lambda row: expected_mapping.get(int(row["NYHA"])) == int(row["label_3class"]), axis=1
        )
        add(
            f"{cohort_name}_nyha_mapping",
            bool(mapping_ok.all()),
            f"mismatch_ids={split.loc[~mapping_ok, 'ID'].astype(str).tolist()}",
        )
        sex_consistency = split.groupby("patient_group_id")["SEX"].nunique().le(1)
        add(
            f"{cohort_name}_group_sex_consistency",
            bool(sex_consistency.all()),
            f"conflicting_groups={sex_consistency[~sex_consistency].index.astype(str).tolist()}",
        )
    return alignment, pd.DataFrame(checks)


def assemble_cohorts(
    exif: pd.DataFrame,
    full_split: pd.DataFrame,
    formal_split: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    exif = exif.copy()
    exif["camera_id"] = exif["Make"].str.strip() + "/" + exif["Model"].str.strip()
    full = full_split.merge(exif, on="ID", how="left", validate="one_to_one")
    formal = formal_split.merge(exif, on="ID", how="left", validate="one_to_one")
    full = derive_parameters(full)
    formal = derive_parameters(formal)
    return {"all_exif_522": full, "formal_global_meanbg_500": formal}


def correlation_pair(x: pd.Series, y: pd.Series) -> dict[str, float]:
    data = pd.DataFrame({"x": pd.to_numeric(x, errors="coerce"), "y": pd.to_numeric(y, errors="coerce")}).dropna()
    data = data.loc[np.isfinite(data["x"]) & np.isfinite(data["y"])]
    if len(data) < 3 or data["x"].nunique() < 2 or data["y"].nunique() < 2:
        return {
            "n": len(data),
            "pearson_r": float("nan"),
            "pearson_p": float("nan"),
            "spearman_rho": float("nan"),
            "spearman_p": float("nan"),
        }
    pearson = stats.pearsonr(data["x"], data["y"])
    spearman = stats.spearmanr(data["x"], data["y"])
    return {
        "n": int(len(data)),
        "pearson_r": float(pearson.statistic),
        "pearson_p": float(pearson.pvalue),
        "spearman_rho": float(spearman.statistic),
        "spearman_p": float(spearman.pvalue),
    }


def apex_consistency(cohorts: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for cohort_name, frame in cohorts.items():
        scopes: list[tuple[str, str, pd.DataFrame]] = [("overall", "ALL", frame)]
        scopes.extend(("within_camera", str(camera), group) for camera, group in frame.groupby("camera_id"))
        for scope, camera_id, subset in scopes:
            data = subset[["brightness_apex_pred", "BrightnessValue", "brightness_residual"]].dropna()
            data = data.loc[np.isfinite(data).all(axis=1)]
            corr = correlation_pair(data["brightness_apex_pred"], data["BrightnessValue"])
            record: dict[str, Any] = {
                "cohort": cohort_name,
                "analysis_type": "fit",
                "scope": scope,
                "camera_id": camera_id,
                **corr,
                "slope": float("nan"),
                "intercept": float("nan"),
                "r_squared": float("nan"),
                "residual_mean": float("nan"),
                "residual_std": float("nan"),
                "residual_median": float("nan"),
                "residual_mad": float("nan"),
            }
            if len(data) >= 3 and data["brightness_apex_pred"].nunique() >= 2:
                fit = stats.linregress(data["brightness_apex_pred"], data["BrightnessValue"])
                record.update(
                    {
                        "slope": float(fit.slope),
                        "intercept": float(fit.intercept),
                        "r_squared": float(fit.rvalue**2),
                        "residual_mean": float(data["brightness_residual"].mean()),
                        "residual_std": float(data["brightness_residual"].std(ddof=1)),
                        "residual_median": float(data["brightness_residual"].median()),
                        "residual_mad": median_absolute_deviation(data["brightness_residual"]),
                    }
                )
            rows.append(record)

        cameras = sorted(frame["camera_id"].dropna().unique().tolist())
        if len(cameras) == 2:
            data = frame[["brightness_apex_pred", "BrightnessValue", "camera_id"]].dropna().copy()
            data = data.loc[np.isfinite(data[["brightness_apex_pred", "BrightnessValue"]]).all(axis=1)]
            data["camera_binary"] = (data["camera_id"] == cameras[1]).astype(float)
            pred = data["brightness_apex_pred"].to_numpy(dtype=float)
            camera = data["camera_binary"].to_numpy(dtype=float)
            design = sm.add_constant(np.column_stack([pred, camera, pred * camera]), has_constant="add")
            model = sm.OLS(data["BrightnessValue"].to_numpy(dtype=float), design).fit()
            names = ["intercept", "brightness_apex_pred", "camera_main", "camera_by_pred_interaction"]
            for index, name in enumerate(names):
                rows.append(
                    {
                        "cohort": cohort_name,
                        "analysis_type": "device_interaction",
                        "scope": "overall",
                        "camera_id": f"reference={cameras[0]};indicator={cameras[1]}",
                        "term": name,
                        "coefficient": float(model.params[index]),
                        "standard_error": float(model.bse[index]),
                        "term_p": float(model.pvalues[index]),
                        "r_squared": float(model.rsquared),
                        "n": int(model.nobs),
                    }
                )
    return pd.DataFrame(rows)


def hedges_g(group_a: np.ndarray, group_b: np.ndarray) -> float:
    a = np.asarray(group_a, dtype=float)
    b = np.asarray(group_b, dtype=float)
    a, b = a[np.isfinite(a)], b[np.isfinite(b)]
    if len(a) < 2 or len(b) < 2:
        return float("nan")
    pooled_numerator = (len(a) - 1) * np.var(a, ddof=1) + (len(b) - 1) * np.var(b, ddof=1)
    pooled_denominator = len(a) + len(b) - 2
    pooled_sd = math.sqrt(max(pooled_numerator / pooled_denominator, 0.0))
    difference = float(np.mean(b) - np.mean(a))
    if pooled_sd == 0:
        return math.copysign(float("inf"), difference) if difference != 0 else 0.0
    correction = 1.0 - 3.0 / (4.0 * (len(a) + len(b)) - 9.0)
    return correction * difference / pooled_sd


def cramers_v(table: pd.DataFrame) -> tuple[float, float]:
    if table.empty or min(table.shape) < 2:
        return float("nan"), float("nan")
    chi2, p_value, _, _ = stats.chi2_contingency(table.to_numpy())
    n = table.to_numpy().sum()
    denominator = n * min(table.shape[0] - 1, table.shape[1] - 1)
    return (math.sqrt(chi2 / denominator) if denominator > 0 else float("nan"), float(p_value))


def align_binary_probabilities(model: Any, probabilities: np.ndarray, positive_class: int = 1) -> np.ndarray:
    classes = list(model.classes_)
    if positive_class not in classes:
        return np.full(len(probabilities), np.nan)
    return probabilities[:, classes.index(positive_class)]


def camera_oof_single_feature(formal: pd.DataFrame, variable: str) -> dict[str, float]:
    cameras = sorted(formal["camera_id"].unique().tolist())
    if len(cameras) != 2:
        raise ValueError(f"Expected two cameras, got {cameras}")
    y = (formal["camera_id"] == cameras[1]).astype(int).to_numpy()
    probabilities = np.full(len(formal), np.nan)
    predictions = np.full(len(formal), -1, dtype=int)
    values = formal[[variable]].to_numpy(dtype=float)
    for fold in sorted(formal["fold"].unique()):
        train_mask = formal["fold"].to_numpy() != fold
        val_mask = ~train_mask
        train_values = values[train_mask].copy()
        val_values = values[val_mask].copy()
        train_median = np.nanmedian(train_values, axis=0)
        train_values = np.where(np.isfinite(train_values), train_values, train_median)
        val_values = np.where(np.isfinite(val_values), val_values, train_median)
        scaler = StandardScaler().fit(train_values)
        train_values = scaler.transform(train_values)
        val_values = scaler.transform(val_values)
        model = LogisticRegression(
            penalty="l2", C=1.0, class_weight="balanced", random_state=SEED, max_iter=2000
        )
        model.fit(train_values, y[train_mask])
        fold_prob = align_binary_probabilities(model, model.predict_proba(val_values), 1)
        probabilities[val_mask] = fold_prob
        predictions[val_mask] = (fold_prob >= 0.5).astype(int)
    return {
        "oof_accuracy": float(accuracy_score(y, predictions)),
        "oof_balanced_accuracy": float(balanced_accuracy_score(y, predictions)),
        "oof_macro_f1": float(f1_score(y, predictions, average="macro")),
        "oof_roc_auc": float(roc_auc_score(y, probabilities)),
    }


def camera_oof_multifeature(formal: pd.DataFrame, variables: Sequence[str]) -> dict[str, float]:
    cameras = sorted(formal["camera_id"].unique().tolist())
    y = (formal["camera_id"] == cameras[1]).astype(int).to_numpy()
    probabilities = np.full(len(formal), np.nan)
    predictions = np.full(len(formal), -1, dtype=int)
    values = formal[list(variables)].to_numpy(dtype=float)
    for fold in sorted(formal["fold"].unique()):
        train_mask = formal["fold"].to_numpy() != fold
        val_mask = ~train_mask
        train_values = values[train_mask].copy()
        val_values = values[val_mask].copy()
        train_median = np.nanmedian(train_values, axis=0)
        train_values = np.where(np.isfinite(train_values), train_values, train_median)
        val_values = np.where(np.isfinite(val_values), val_values, train_median)
        scaler = StandardScaler().fit(train_values)
        model = LogisticRegression(
            penalty="l2", C=1.0, class_weight="balanced", random_state=SEED, max_iter=3000
        )
        model.fit(scaler.transform(train_values), y[train_mask])
        fold_prob = align_binary_probabilities(model, model.predict_proba(scaler.transform(val_values)), 1)
        probabilities[val_mask] = fold_prob
        predictions[val_mask] = (fold_prob >= 0.5).astype(int)
    return {
        "oof_accuracy": float(accuracy_score(y, predictions)),
        "oof_balanced_accuracy": float(balanced_accuracy_score(y, predictions)),
        "oof_macro_f1": float(f1_score(y, predictions, average="macro")),
        "oof_roc_auc": float(roc_auc_score(y, probabilities)),
    }


def device_identity_analysis(formal: pd.DataFrame) -> pd.DataFrame:
    cameras = sorted(formal["camera_id"].unique().tolist())
    if len(cameras) != 2:
        raise ValueError(f"Device identity analysis requires two cameras, got {cameras}")
    y = (formal["camera_id"] == cameras[1]).astype(int).to_numpy()
    rows = []
    for variable in ANALYSIS_FIELDS:
        values = formal[variable].to_numpy(dtype=float)
        finite = np.isfinite(values)
        mi = float("nan")
        if finite.sum() >= 10 and np.unique(values[finite]).size >= 2:
            mi = float(
                mutual_info_classif(
                    values[finite].reshape(-1, 1),
                    y[finite],
                    discrete_features=False,
                    random_state=SEED,
                )[0]
            )
        record = {
            "feature": variable,
            "analysis": "single_field",
            "camera_a": cameras[0],
            "camera_b": cameras[1],
            "hedges_g_b_minus_a": hedges_g(values[y == 0], values[y == 1]),
            "mutual_information": mi,
        }
        record.update(camera_oof_single_feature(formal, variable))
        rows.append(record)
    combined_variables = CORE_FIELDS + DEVICE_MODEL_DERIVED
    combined = {
        "feature": "+".join(combined_variables),
        "analysis": "combined_core_and_derived_logistic",
        "camera_a": cameras[0],
        "camera_b": cameras[1],
        "hedges_g_b_minus_a": float("nan"),
        "mutual_information": float("nan"),
    }
    combined.update(camera_oof_multifeature(formal, combined_variables))
    rows.append(combined)
    return pd.DataFrame(rows)


def auxiliary_device_analysis(formal: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    association_rows = []
    crosstab_rows = []
    for field in ["MeteringMode", "Flash"]:
        table = pd.crosstab(formal[field], formal["camera_id"])
        effect, p_value = cramers_v(table)
        association_rows.append(
            {
                "auxiliary_field": field,
                "analysis": "camera_association",
                "cramers_v": effect,
                "chi_square_p": p_value,
                "unique_n": int(formal[field].nunique()),
            }
        )
        for value, row in table.iterrows():
            for camera_id, count in row.items():
                crosstab_rows.append(
                    {"auxiliary_field": field, "value": value, "camera_id": camera_id, "count": int(count)}
                )

    aux = formal[["MeteringMode", "Flash"]].astype(str)
    camera = formal[["camera_id"]].astype(str)
    encoder_aux = OneHotEncoder(handle_unknown="ignore", sparse_output=False).fit(aux)
    encoder_camera = OneHotEncoder(handle_unknown="ignore", sparse_output=False, drop="first").fit(camera)
    x_aux = encoder_aux.transform(aux)
    x_camera = encoder_camera.transform(camera)
    for variable in CORE_FIELDS:
        y = formal[variable].to_numpy(dtype=float)
        for name, design in [
            ("auxiliary_only", x_aux),
            ("camera_only", x_camera),
            ("auxiliary_plus_camera", np.column_stack([x_aux, x_camera])),
        ]:
            r2 = LinearRegression().fit(design, y).score(design, y)
            association_rows.append(
                {
                    "auxiliary_field": "MeteringMode+Flash",
                    "analysis": name,
                    "core_field": variable,
                    "r_squared": float(r2),
                }
            )
    return pd.DataFrame(association_rows), pd.DataFrame(crosstab_rows)


def empirical_overlap(groups: Sequence[np.ndarray]) -> float:
    arrays = [np.asarray(group, dtype=float) for group in groups]
    arrays = [array[np.isfinite(array)] for array in arrays if np.isfinite(array).sum() > 0]
    if len(arrays) < 2:
        return float("nan")
    combined = np.concatenate(arrays)
    if np.ptp(combined) == 0:
        return 1.0
    bins = np.histogram_bin_edges(combined, bins="fd")
    if len(bins) < 3:
        bins = np.linspace(float(combined.min()), float(combined.max()), 11)
    histograms = []
    for array in arrays:
        counts, _ = np.histogram(array, bins=bins)
        histograms.append(counts / max(counts.sum(), 1))
    overlaps = []
    for i in range(len(histograms)):
        for j in range(i + 1, len(histograms)):
            overlaps.append(float(np.minimum(histograms[i], histograms[j]).sum()))
    return min(overlaps) if overlaps else float("nan")


def bh_fdr(p_values: pd.Series) -> pd.Series:
    result = pd.Series(np.nan, index=p_values.index, dtype=float)
    valid = p_values.dropna().astype(float)
    if valid.empty:
        return result
    order = valid.sort_values().index
    ranked = valid.loc[order].to_numpy()
    adjusted = ranked * len(ranked) / np.arange(1, len(ranked) + 1)
    adjusted = np.minimum.accumulate(adjusted[::-1])[::-1]
    adjusted = np.clip(adjusted, 0, 1)
    result.loc[order] = adjusted
    return result


def confounding_tests(cohorts: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for cohort_name, frame in cohorts.items():
        contexts: list[tuple[str, str, pd.DataFrame, bool]] = [("raw_overall", "ALL", frame, False)]
        contexts.extend(
            ("within_camera", str(camera_id), subset, False)
            for camera_id, subset in frame.groupby("camera_id")
        )
        contexts.append(("device_adjusted", "ALL", frame, True))
        for context, camera_id, subset, adjust in contexts:
            for variable in ANALYSIS_FIELDS:
                values = pd.to_numeric(subset[variable], errors="coerce").astype(float)
                if adjust:
                    adjusted = pd.Series(index=subset.index, dtype=float)
                    for _, indices in subset.groupby("camera_id").groups.items():
                        camera_values = values.loc[indices]
                        adjusted.loc[indices] = camera_values - camera_values.median()
                    values = adjusted

                nyha_groups = [
                    values.loc[subset["label_3class"] == label].dropna().to_numpy(dtype=float)
                    for label in [0, 1, 2]
                ]
                if all(len(group) > 0 for group in nyha_groups) and np.unique(np.concatenate(nyha_groups)).size > 1:
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        test = stats.kruskal(*nyha_groups)
                    n = sum(len(group) for group in nyha_groups)
                    epsilon_sq = max(0.0, (float(test.statistic) - 3 + 1) / max(n - 3, 1))
                    p_value = float(test.pvalue)
                    statistic = float(test.statistic)
                else:
                    n = sum(len(group) for group in nyha_groups)
                    epsilon_sq = float("nan")
                    p_value = float("nan")
                    statistic = float("nan")
                rows.append(
                    {
                        "cohort": cohort_name,
                        "outcome": "NYHA_3class",
                        "comparison_context": context,
                        "camera_id": camera_id,
                        "variable": variable,
                        "test": "Kruskal-Wallis",
                        "n": n,
                        "statistic": statistic,
                        "p_value": p_value,
                        "effect_name": "epsilon_squared",
                        "effect_size": epsilon_sq,
                        "distribution_overlap_min": empirical_overlap(nyha_groups),
                    }
                )

                valid = pd.DataFrame({"value": values, "label": subset["label_3class"]}).dropna()
                if len(valid) >= 3 and valid["value"].nunique() > 1 and valid["label"].nunique() > 1:
                    correlation = stats.spearmanr(valid["value"], valid["label"])
                    rows.append(
                        {
                            "cohort": cohort_name,
                            "outcome": "NYHA_ordered_exploratory",
                            "comparison_context": context,
                            "camera_id": camera_id,
                            "variable": variable,
                            "test": "Spearman",
                            "n": len(valid),
                            "statistic": float(correlation.statistic),
                            "p_value": float(correlation.pvalue),
                            "effect_name": "spearman_rho",
                            "effect_size": float(correlation.statistic),
                            "distribution_overlap_min": float("nan"),
                        }
                    )

                female = values.loc[subset["SEX"] == 0].dropna().to_numpy(dtype=float)
                male = values.loc[subset["SEX"] == 1].dropna().to_numpy(dtype=float)
                if len(female) > 0 and len(male) > 0 and np.unique(np.concatenate([female, male])).size > 1:
                    test = stats.mannwhitneyu(male, female, alternative="two-sided")
                    rank_biserial = 2.0 * float(test.statistic) / (len(male) * len(female)) - 1.0
                    p_value = float(test.pvalue)
                    statistic = float(test.statistic)
                else:
                    rank_biserial = float("nan")
                    p_value = float("nan")
                    statistic = float("nan")
                rows.append(
                    {
                        "cohort": cohort_name,
                        "outcome": "SEX_male_vs_female",
                        "comparison_context": context,
                        "camera_id": camera_id,
                        "variable": variable,
                        "test": "Mann-Whitney U",
                        "n": len(male) + len(female),
                        "statistic": statistic,
                        "p_value": p_value,
                        "effect_name": "rank_biserial_male_minus_female",
                        "effect_size": rank_biserial,
                        "distribution_overlap_min": empirical_overlap([male, female]),
                    }
                )
    output = pd.DataFrame(rows)
    output["p_fdr_bh"] = np.nan
    family = ["cohort", "outcome", "comparison_context", "camera_id", "test"]
    for _, indices in output.groupby(family, dropna=False).groups.items():
        output.loc[indices, "p_fdr_bh"] = bh_fdr(output.loc[indices, "p_value"])
    return output


def compute_image_quality(
    frame: pd.DataFrame,
    aligned_dir: Path,
    parsing_label_dir: Path,
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    rows = []
    failures = []
    for position, image in enumerate(frame.itertuples(index=False), start=1):
        image_id = str(image.ID)
        rgb_path = aligned_dir / f"{image_id}.png"
        label_path = parsing_label_dir / f"{image_id}.png"
        try:
            rgb_u8 = np.asarray(Image.open(rgb_path).convert("RGB"), dtype=np.uint8)
            label = np.asarray(Image.open(label_path), dtype=np.uint8)
            if label.shape != rgb_u8.shape[:2]:
                raise ValueError(f"shape mismatch rgb={rgb_u8.shape}, label={label.shape}")
            skin = label == 1
            skin_n = int(skin.sum())
            if skin_n < 64:
                raise ValueError(f"skin mask too small: {skin_n}")
            inner = ndimage.binary_erosion(skin, iterations=2)
            if int(inner.sum()) < 64:
                inner = skin

            rgb = rgb_u8.astype(np.float64) / 255.0
            pixels = rgb[skin]
            linear = np.where(
                rgb <= 0.04045,
                rgb / 12.92,
                ((rgb + 0.055) / 1.055) ** 2.4,
            )
            luminance = 0.2126 * linear[..., 0] + 0.7152 * linear[..., 1] + 0.0722 * linear[..., 2]
            skin_luminance = luminance[skin]
            delta = 6.0 / 29.0
            f_y = np.where(
                skin_luminance > delta**3,
                np.cbrt(skin_luminance),
                skin_luminance / (3 * delta**2) + 4.0 / 29.0,
            )
            lab_l = 116.0 * f_y - 16.0
            gray = 0.299 * rgb[..., 0] + 0.587 * rgb[..., 1] + 0.114 * rgb[..., 2]
            laplacian = ndimage.laplace(gray, mode="reflect")
            smooth = ndimage.gaussian_filter(gray, sigma=1.0, mode="reflect")
            residual = (gray - smooth)[inner]
            residual_median = float(np.median(residual))

            rows.append(
                {
                    "ID": image_id,
                    "skin_pixel_n": skin_n,
                    "valid_skin_pixel_ratio": skin_n / float(skin.size),
                    "skin_median_r": float(np.median(pixels[:, 0])),
                    "skin_median_g": float(np.median(pixels[:, 1])),
                    "skin_median_b": float(np.median(pixels[:, 2])),
                    "skin_linear_luminance_median": float(np.median(skin_luminance)),
                    "skin_lab_l_median": float(np.median(lab_l)),
                    "overexposed_pixel_ratio": float(np.mean(skin_luminance >= 0.98)),
                    "underexposed_pixel_ratio": float(np.mean(skin_luminance <= 0.02)),
                    "saturated_pixel_ratio": float(
                        np.mean(np.any((pixels <= 5.0 / 255.0) | (pixels >= 250.0 / 255.0), axis=1))
                    ),
                    "laplacian_variance": float(np.var(laplacian[inner], ddof=1)),
                    "high_frequency_noise_mad": float(
                        1.4826 * np.median(np.abs(residual - residual_median))
                    ),
                }
            )
        except Exception as exc:
            failures.append({"ID": image_id, "error": f"{type(exc).__name__}: {exc}"})
        if position % 100 == 0:
            print(f"[image-quality] processed {position}/{len(frame)}", flush=True)
    return pd.DataFrame(rows), failures


def image_quality_correlations(frame: pd.DataFrame) -> pd.DataFrame:
    pairs = [
        ("ExposureTime", "skin_linear_luminance_median"),
        ("ExposureTime", "laplacian_variance"),
        ("ExposureTime", "high_frequency_noise_mad"),
        ("ISOSpeedRatings", "high_frequency_noise_mad"),
        ("ISOSpeedRatings", "underexposed_pixel_ratio"),
        ("ISOSpeedRatings", "skin_linear_luminance_median"),
        ("FNumber", "skin_linear_luminance_median"),
        ("FNumber", "laplacian_variance"),
        ("BrightnessValue", "skin_linear_luminance_median"),
        ("BrightnessValue", "skin_lab_l_median"),
        ("BrightnessValue", "overexposed_pixel_ratio"),
        ("BrightnessValue", "underexposed_pixel_ratio"),
    ]
    rows = []
    scopes: list[tuple[str, str, pd.DataFrame]] = [("overall", "ALL", frame)]
    scopes.extend(("within_camera", str(camera), group) for camera, group in frame.groupby("camera_id"))
    for scope, camera_id, subset in scopes:
        for exif_field, image_metric in pairs:
            result = correlation_pair(subset[exif_field], subset[image_metric])
            rows.append(
                {
                    "scope": scope,
                    "camera_id": camera_id,
                    "exif_field": exif_field,
                    "image_metric": image_metric,
                    **result,
                }
            )
    return pd.DataFrame(rows)


FEATURE_GROUPS: dict[str, dict[str, Any]] = {
    "E0_log2_exposure_time": {"numeric": ["log2_exposure_time"], "camera": False},
    "E0_log2_iso_gain": {"numeric": ["log2_iso_gain"], "camera": False},
    "E0_log2_fnumber": {"numeric": ["log2_fnumber"], "camera": False},
    "E0_BrightnessValue": {"numeric": ["BrightnessValue"], "camera": False},
    "E1_four_core_transformed": {
        "numeric": ["log2_exposure_time", "log2_iso_gain", "log2_fnumber", "BrightnessValue"],
        "camera": False,
    },
    "E2_compact_physical": {
        "numeric": ["relative_optical_exposure", "log2_iso_gain", "BrightnessValue"],
        "camera": False,
    },
    "E3_device_centered_physical": {
        "numeric": ["relative_optical_exposure", "log2_iso_gain", "device_centered_brightness_fold"],
        "camera": False,
    },
    "E4_camera_only": {"numeric": [], "camera": True},
    "E5_E3_plus_camera": {
        "numeric": ["relative_optical_exposure", "log2_iso_gain", "device_centered_brightness_fold"],
        "camera": True,
    },
}


def fold_centered_brightness(
    train: pd.DataFrame,
    target: pd.DataFrame,
) -> tuple[np.ndarray, dict[str, dict[str, float]]]:
    output = np.full(len(target), np.nan, dtype=float)
    fallback_median = float(train["BrightnessValue"].median())
    fallback_iqr = float(train["BrightnessValue"].quantile(0.75) - train["BrightnessValue"].quantile(0.25))
    if not math.isfinite(fallback_iqr) or fallback_iqr <= 0:
        fallback_iqr = 1.0
    stats_by_device: dict[str, dict[str, float]] = {}
    for camera_id, group in train.groupby("camera_id"):
        median = float(group["BrightnessValue"].median())
        iqr = float(group["BrightnessValue"].quantile(0.75) - group["BrightnessValue"].quantile(0.25))
        if not math.isfinite(iqr) or iqr <= 0:
            iqr = 1.0
        stats_by_device[str(camera_id)] = {"median": median, "iqr": iqr}
    for position, (_, row) in enumerate(target.iterrows()):
        item = stats_by_device.get(str(row["camera_id"]), {"median": fallback_median, "iqr": fallback_iqr})
        output[position] = (float(row["BrightnessValue"]) - item["median"]) / item["iqr"]
    return output, stats_by_device


def build_fold_features(
    train: pd.DataFrame,
    validation: pd.DataFrame,
    feature_spec: dict[str, Any],
) -> tuple[np.ndarray, np.ndarray, dict[str, Any]]:
    train_work = train.copy()
    validation_work = validation.copy()
    preprocessing: dict[str, Any] = {}
    numeric_fields = list(feature_spec["numeric"])
    if "device_centered_brightness_fold" in numeric_fields:
        train_centered, device_stats = fold_centered_brightness(train_work, train_work)
        validation_centered, _ = fold_centered_brightness(train_work, validation_work)
        train_work["device_centered_brightness_fold"] = train_centered
        validation_work["device_centered_brightness_fold"] = validation_centered
        preprocessing["device_brightness_train_median_iqr"] = device_stats

    blocks_train: list[np.ndarray] = []
    blocks_val: list[np.ndarray] = []
    if numeric_fields:
        x_train = train_work[numeric_fields].to_numpy(dtype=float)
        x_val = validation_work[numeric_fields].to_numpy(dtype=float)
        medians = np.nanmedian(x_train, axis=0)
        medians = np.where(np.isfinite(medians), medians, 0.0)
        x_train = np.where(np.isfinite(x_train), x_train, medians)
        x_val = np.where(np.isfinite(x_val), x_val, medians)
        scaler = StandardScaler().fit(x_train)
        blocks_train.append(scaler.transform(x_train))
        blocks_val.append(scaler.transform(x_val))
        preprocessing["numeric_fields"] = numeric_fields
        preprocessing["train_imputation_medians"] = medians.tolist()
        preprocessing["train_scaler_mean"] = scaler.mean_.tolist()
        preprocessing["train_scaler_scale"] = scaler.scale_.tolist()

    if feature_spec["camera"]:
        encoder = OneHotEncoder(handle_unknown="ignore", sparse_output=False)
        train_camera = train_work[["camera_id"]].astype(str)
        val_camera = validation_work[["camera_id"]].astype(str)
        blocks_train.append(encoder.fit_transform(train_camera))
        blocks_val.append(encoder.transform(val_camera))
        preprocessing["camera_categories"] = [list(map(str, values)) for values in encoder.categories_]

    if not blocks_train:
        raise ValueError("Feature group produced no features")
    return np.column_stack(blocks_train), np.column_stack(blocks_val), preprocessing


def multiclass_metrics(y_true: np.ndarray, probabilities: np.ndarray) -> dict[str, float]:
    y_true = np.asarray(y_true, dtype=int)
    probabilities = np.asarray(probabilities, dtype=float)
    predictions = probabilities.argmax(axis=1)
    metrics: dict[str, float] = {
        "accuracy": float(accuracy_score(y_true, predictions)),
        "balanced_accuracy": float(balanced_accuracy_score(y_true, predictions)),
        "macro_precision": float(precision_score(y_true, predictions, average="macro", zero_division=0)),
        "macro_recall": float(recall_score(y_true, predictions, average="macro", zero_division=0)),
        "macro_f1": float(f1_score(y_true, predictions, average="macro", zero_division=0)),
    }
    for label, name in CLASS_NAMES.items():
        target = (y_true == label).astype(int)
        metrics[f"{name}_auc"] = (
            float(roc_auc_score(target, probabilities[:, label])) if np.unique(target).size == 2 else float("nan")
        )
    class_auc = [metrics[f"{name}_auc"] for name in CLASS_NAMES.values()]
    metrics["macro_auc"] = float(np.nanmean(class_auc))
    return metrics


def align_multiclass_probabilities(model: Any, probabilities: np.ndarray) -> np.ndarray:
    output = np.zeros((len(probabilities), 3), dtype=float)
    for column, label in enumerate(model.classes_):
        output[:, int(label)] = probabilities[:, column]
    return output


def exif_only_oof(formal: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    prediction_rows: list[pd.DataFrame] = []
    fold_metric_rows: list[dict[str, Any]] = []
    confusion_rows: list[dict[str, Any]] = []
    preprocessing_audit: dict[str, Any] = {}
    model_types = ["multinomial_logistic", "random_forest_shallow"]

    for model_type in model_types:
        for feature_group, feature_spec in FEATURE_GROUPS.items():
            oof = np.full((len(formal), 3), np.nan, dtype=float)
            preprocessing_audit[f"{model_type}/{feature_group}"] = {}
            for fold in sorted(formal["fold"].unique()):
                train = formal.loc[formal["fold"] != fold].copy()
                validation = formal.loc[formal["fold"] == fold].copy()
                x_train, x_val, preprocessing = build_fold_features(train, validation, feature_spec)
                y_train = train["label_3class"].to_numpy(dtype=int)
                if model_type == "multinomial_logistic":
                    model = LogisticRegression(
                        penalty="l2",
                        C=1.0,
                        solver="lbfgs",
                        class_weight="balanced",
                        random_state=SEED,
                        max_iter=3000,
                    )
                else:
                    model = RandomForestClassifier(
                        n_estimators=300,
                        max_depth=3,
                        min_samples_leaf=20,
                        class_weight="balanced",
                        random_state=SEED,
                        n_jobs=1,
                    )
                model.fit(x_train, y_train)
                probabilities = align_multiclass_probabilities(model, model.predict_proba(x_val))
                oof[validation.index.to_numpy()] = probabilities
                metrics = multiclass_metrics(validation["label_3class"].to_numpy(dtype=int), probabilities)
                for metric_name, metric_value in metrics.items():
                    fold_metric_rows.append(
                        {
                            "model_type": model_type,
                            "feature_group": feature_group,
                            "fold": int(fold),
                            "metric": metric_name,
                            "value": metric_value,
                            "train_n": len(train),
                            "validation_n": len(validation),
                            "train_patient_group_n": train["patient_group_id"].nunique(),
                            "validation_patient_group_n": validation["patient_group_id"].nunique(),
                        }
                    )
                preprocessing_audit[f"{model_type}/{feature_group}"][str(fold)] = preprocessing

            if not np.isfinite(oof).all():
                raise RuntimeError(f"Non-finite OOF probabilities for {model_type}/{feature_group}")
            row_sums = oof.sum(axis=1)
            if not np.allclose(row_sums, 1.0, atol=1e-8):
                raise RuntimeError(f"OOF probability rows do not sum to one for {model_type}/{feature_group}")
            predictions = oof.argmax(axis=1)
            prediction_rows.append(
                pd.DataFrame(
                    {
                        "ID": formal["ID"].astype(str),
                        "patient_group_id": formal["patient_group_id"].astype(str),
                        "fold": formal["fold"].astype(int),
                        "NYHA": formal["NYHA"].astype(int),
                        "SEX": formal["SEX"].astype(int),
                        "camera_id": formal["camera_id"].astype(str),
                        "label_3class": formal["label_3class"].astype(int),
                        "label_3class_name": formal["label_3class_name"].astype(str),
                        "model_type": model_type,
                        "feature_group": feature_group,
                        "prob_normal": oof[:, 0],
                        "prob_mild": oof[:, 1],
                        "prob_severe": oof[:, 2],
                        "predicted_label": predictions,
                        "predicted_name": pd.Series(predictions).map(CLASS_NAMES),
                    }
                )
            )
            matrix = confusion_matrix(formal["label_3class"].to_numpy(dtype=int), predictions, labels=[0, 1, 2])
            for true_label in range(3):
                for predicted_label in range(3):
                    confusion_rows.append(
                        {
                            "model_type": model_type,
                            "feature_group": feature_group,
                            "true_label": true_label,
                            "true_name": CLASS_NAMES[true_label],
                            "predicted_label": predicted_label,
                            "predicted_name": CLASS_NAMES[predicted_label],
                            "count": int(matrix[true_label, predicted_label]),
                        }
                    )

    predictions = pd.concat(prediction_rows, ignore_index=True)
    fold_metrics = pd.DataFrame(fold_metric_rows)
    confusion = pd.DataFrame(confusion_rows)

    aggregate_rows = []
    for (model_type, feature_group), subset in predictions.groupby(["model_type", "feature_group"], sort=True):
        y_true = subset["label_3class"].to_numpy(dtype=int)
        probabilities = subset[["prob_normal", "prob_mild", "prob_severe"]].to_numpy(dtype=float)
        point = multiclass_metrics(y_true, probabilities)
        groups = subset["patient_group_id"].astype(str).to_numpy()
        unique_groups = np.unique(groups)
        group_indices = {group: np.flatnonzero(groups == group) for group in unique_groups}
        rng = np.random.default_rng(SEED)
        bootstrap_values = {metric: [] for metric in point}
        for _ in range(BOOTSTRAP_REPEATS):
            sampled_groups = rng.choice(unique_groups, size=len(unique_groups), replace=True)
            sampled_indices = np.concatenate([group_indices[group] for group in sampled_groups])
            sample_metrics = multiclass_metrics(y_true[sampled_indices], probabilities[sampled_indices])
            for metric_name, value in sample_metrics.items():
                if math.isfinite(value):
                    bootstrap_values[metric_name].append(value)
        for metric_name, point_value in point.items():
            fold_values = fold_metrics.loc[
                (fold_metrics["model_type"] == model_type)
                & (fold_metrics["feature_group"] == feature_group)
                & (fold_metrics["metric"] == metric_name),
                "value",
            ]
            bootstrap = np.asarray(bootstrap_values[metric_name], dtype=float)
            aggregate_rows.append(
                {
                    "model_type": model_type,
                    "feature_group": feature_group,
                    "metric": metric_name,
                    "oof_value": point_value,
                    "fold_mean": float(fold_values.mean()),
                    "fold_std": float(fold_values.std(ddof=1)),
                    "bootstrap_ci_low_95": float(np.quantile(bootstrap, 0.025)) if len(bootstrap) else float("nan"),
                    "bootstrap_ci_high_95": float(np.quantile(bootstrap, 0.975)) if len(bootstrap) else float("nan"),
                    "bootstrap_valid_repeats": int(len(bootstrap)),
                    "bootstrap_requested_repeats": BOOTSTRAP_REPEATS,
                    "bootstrap_seed": SEED,
                    "bootstrap_unit": "patient_group_id",
                }
            )
        print(f"[bootstrap] completed {model_type}/{feature_group}", flush=True)
    aggregate = pd.DataFrame(aggregate_rows)

    validation_rows = []
    for (model_type, feature_group), subset in predictions.groupby(["model_type", "feature_group"]):
        probability_array = subset[["prob_normal", "prob_mild", "prob_severe"]].to_numpy(dtype=float)
        validation_rows.append(
            {
                "model_type": model_type,
                "feature_group": feature_group,
                "rows": len(subset),
                "unique_ids": subset["ID"].nunique(),
                "each_id_once": len(subset) == subset["ID"].nunique() == len(formal),
                "probabilities_finite": bool(np.isfinite(probability_array).all()),
                "max_probability_sum_abs_error": float(np.abs(probability_array.sum(axis=1) - 1).max()),
                "same_patient_cross_fold": bool(subset.groupby("patient_group_id")["fold"].nunique().gt(1).any()),
            }
        )
    validation = pd.DataFrame(validation_rows)
    return predictions, aggregate, fold_metrics, confusion, {
        "fold_preprocessing": preprocessing_audit,
        "validation": validation.to_dict(orient="records"),
    }


def correlation_tables(cohorts: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for cohort_name, frame in cohorts.items():
        scopes: list[tuple[str, pd.DataFrame]] = [("overall", frame)]
        scopes.extend((str(camera), group) for camera, group in frame.groupby("camera_id"))
        for scope, subset in scopes:
            for method in ["pearson", "spearman"]:
                matrix = subset[ANALYSIS_FIELDS].corr(method=method)
                for variable_a in ANALYSIS_FIELDS:
                    for variable_b in ANALYSIS_FIELDS:
                        rows.append(
                            {
                                "cohort": cohort_name,
                                "scope": scope,
                                "method": method,
                                "variable_a": variable_a,
                                "variable_b": variable_b,
                                "correlation": matrix.loc[variable_a, variable_b],
                                "n_pairwise": int(subset[[variable_a, variable_b]].dropna().shape[0]),
                            }
                        )
    return pd.DataFrame(rows)


def calculate_vif(frame: pd.DataFrame, cohort: str, scope: str) -> pd.DataFrame:
    rows = []
    work = frame[ANALYSIS_FIELDS].replace([np.inf, -np.inf], np.nan).dropna()
    for variable in ANALYSIS_FIELDS:
        y = work[variable].to_numpy(dtype=float)
        if len(y) < 5 or np.unique(y).size < 2:
            vif = float("nan")
            reason = "zero_variance_or_insufficient"
        else:
            predictors = [field for field in ANALYSIS_FIELDS if field != variable and work[field].nunique() > 1]
            if not predictors:
                vif = float("nan")
                reason = "no_nonconstant_predictors"
            else:
                r2 = LinearRegression().fit(work[predictors].to_numpy(dtype=float), y).score(
                    work[predictors].to_numpy(dtype=float), y
                )
                vif = float("inf") if r2 >= 1.0 - 1e-12 else float(1.0 / max(1.0 - r2, 1e-12))
                reason = "exact_or_near_exact_dependency" if math.isinf(vif) else "estimated"
        rows.append(
            {
                "cohort": cohort,
                "scope": scope,
                "variable": variable,
                "vif": vif,
                "reason": reason,
                "n_complete": len(work),
            }
        )
    return pd.DataFrame(rows)


def vif_tables(cohorts: dict[str, pd.DataFrame]) -> pd.DataFrame:
    outputs = []
    for cohort_name, frame in cohorts.items():
        outputs.append(calculate_vif(frame, cohort_name, "overall"))
        for camera_id, subset in frame.groupby("camera_id"):
            outputs.append(calculate_vif(subset, cohort_name, str(camera_id)))
    return pd.concat(outputs, ignore_index=True)


def plot_core_distributions(formal: pd.DataFrame, figures_dir: Path) -> None:
    fig, axes = plt.subplots(2, 2, figsize=(7.2, 5.4))
    for ax, variable in zip(axes.flat, CORE_FIELDS):
        for camera_id, group in formal.groupby("camera_id"):
            ax.hist(
                group[variable].dropna(),
                bins=24,
                density=True,
                alpha=0.55,
                color=PALETTE.get(str(camera_id), "#777777"),
                label=f"{camera_id} (n={len(group)})",
            )
        ax.set_title(variable, fontweight="bold")
        ax.set_xlabel(variable)
        ax.set_ylabel("Density")
        ax.grid(axis="y", alpha=0.18)
    axes.flat[0].legend(fontsize=6.5)
    fig.suptitle("Core EXIF distributions by camera", fontsize=10, fontweight="bold")
    fig.tight_layout()
    save_figure(fig, figures_dir / "core_original_distributions_by_device")


def plot_core_by_nyha(formal: pd.DataFrame, figures_dir: Path) -> None:
    fig, axes = plt.subplots(2, 2, figsize=(7.2, 5.4))
    for ax, variable in zip(axes.flat, CORE_FIELDS):
        data = [formal.loc[formal["label_3class"] == label, variable].dropna().to_numpy() for label in range(3)]
        box = ax.boxplot(data, positions=[0, 1, 2], widths=0.55, patch_artist=True, showfliers=False)
        for patch, name in zip(box["boxes"], CLASS_NAMES.values()):
            patch.set_facecolor(PALETTE[name])
            patch.set_alpha(0.72)
        ax.set_xticks([0, 1, 2], [CLASS_NAMES[index] for index in range(3)])
        ax.set_title(variable, fontweight="bold")
        ax.set_ylabel(variable)
        ax.grid(axis="y", alpha=0.18)
    fig.suptitle("Core EXIF parameters across NYHA groups", fontsize=10, fontweight="bold")
    fig.tight_layout()
    save_figure(fig, figures_dir / "core_parameters_by_nyha")


def plot_camera_nyha(formal: pd.DataFrame, figures_dir: Path) -> None:
    cameras = sorted(formal["camera_id"].unique().tolist())
    fig, axes = plt.subplots(2, 2, figsize=(7.2, 5.6))
    for ax, variable in zip(axes.flat, CORE_FIELDS):
        data = []
        positions = []
        colors = []
        for label in range(3):
            for camera_index, camera_id in enumerate(cameras):
                values = formal.loc[
                    (formal["label_3class"] == label) & (formal["camera_id"] == camera_id), variable
                ].dropna().to_numpy()
                data.append(values)
                positions.append(label + (-0.18 if camera_index == 0 else 0.18))
                colors.append(PALETTE.get(camera_id, "#777777"))
        box = ax.boxplot(data, positions=positions, widths=0.28, patch_artist=True, showfliers=False)
        for patch, color in zip(box["boxes"], colors):
            patch.set_facecolor(color)
            patch.set_alpha(0.72)
        ax.set_xticks([0, 1, 2], [CLASS_NAMES[index] for index in range(3)])
        ax.set_title(variable, fontweight="bold")
        ax.set_ylabel(variable)
        ax.grid(axis="y", alpha=0.18)
    handles = [plt.Line2D([0], [0], color=PALETTE.get(camera, "#777777"), lw=6) for camera in cameras]
    fig.legend(handles, cameras, loc="upper center", ncol=2, fontsize=7)
    fig.suptitle("Camera-stratified NYHA distributions", y=0.98, fontsize=10, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.93])
    save_figure(fig, figures_dir / "camera_by_nyha_grouped")


def plot_correlation_heatmap(formal: pd.DataFrame, figures_dir: Path) -> None:
    matrix = formal[ANALYSIS_FIELDS].corr(method="spearman")
    safe = matrix.fillna(0.0).clip(-1, 1)
    distance = 1.0 - np.abs(safe.to_numpy())
    np.fill_diagonal(distance, 0.0)
    condensed = squareform(distance, checks=False)
    linkage = hierarchy.linkage(condensed, method="average")
    order = hierarchy.leaves_list(linkage)
    ordered = matrix.iloc[order, order]
    fig, ax = plt.subplots(figsize=(7.2, 6.6))
    image = ax.imshow(ordered, cmap="coolwarm", vmin=-1, vmax=1, aspect="equal")
    labels = [ordered.index[index] for index in range(len(ordered))]
    ax.set_xticks(range(len(labels)), labels, rotation=60, ha="right", fontsize=6)
    ax.set_yticks(range(len(labels)), labels, fontsize=6)
    ax.set_title("Spearman correlation of core and derived EXIF variables", fontweight="bold")
    colorbar = fig.colorbar(image, ax=ax, fraction=0.04, pad=0.03)
    colorbar.set_label("Spearman rho")
    fig.tight_layout()
    save_figure(fig, figures_dir / "core_and_derived_correlation_heatmap")


def plot_apex(formal: pd.DataFrame, figures_dir: Path) -> None:
    fig, ax = plt.subplots(figsize=(5.5, 4.3))
    for camera_id, group in formal.groupby("camera_id"):
        color = PALETTE.get(str(camera_id), "#777777")
        ax.scatter(
            group["brightness_apex_pred"],
            group["BrightnessValue"],
            s=13,
            alpha=0.55,
            color=color,
            label=f"{camera_id} (n={len(group)})",
        )
        x = group["brightness_apex_pred"].to_numpy(dtype=float)
        y = group["BrightnessValue"].to_numpy(dtype=float)
        slope, intercept = np.polyfit(x, y, 1)
        grid = np.linspace(np.nanmin(x), np.nanmax(x), 100)
        ax.plot(grid, slope * grid + intercept, color=color, lw=1.5)
    limits = [
        min(formal["brightness_apex_pred"].min(), formal["BrightnessValue"].min()),
        max(formal["brightness_apex_pred"].max(), formal["BrightnessValue"].max()),
    ]
    ax.plot(limits, limits, ls="--", color="#555555", lw=1, label="identity")
    ax.set_xlabel("APEX-predicted BrightnessValue")
    ax.set_ylabel("Recorded BrightnessValue")
    ax.set_title("Recorded versus APEX-predicted brightness", fontweight="bold")
    ax.legend(fontsize=6.5)
    ax.grid(alpha=0.18)
    fig.tight_layout()
    save_figure(fig, figures_dir / "brightness_value_vs_apex_prediction")

    fig, ax = plt.subplots(figsize=(5.2, 4.0))
    cameras = sorted(formal["camera_id"].unique().tolist())
    data = [formal.loc[formal["camera_id"] == camera, "brightness_residual"].to_numpy() for camera in cameras]
    box = ax.boxplot(data, labels=cameras, patch_artist=True, showfliers=True, flierprops={"markersize": 2})
    for patch, camera in zip(box["boxes"], cameras):
        patch.set_facecolor(PALETTE.get(camera, "#777777"))
        patch.set_alpha(0.72)
    ax.axhline(0, color="#555555", lw=1, ls="--")
    ax.set_ylabel("Brightness residual (EV)")
    ax.set_title("APEX brightness residual by camera", fontweight="bold")
    ax.tick_params(axis="x", rotation=15)
    ax.grid(axis="y", alpha=0.18)
    fig.tight_layout()
    save_figure(fig, figures_dir / "brightness_residual_by_device")


def plot_image_quality_relations(formal: pd.DataFrame, figures_dir: Path) -> None:
    pairs = [
        ("ISOSpeedRatings", "high_frequency_noise_mad", "ISO versus high-frequency noise"),
        ("ExposureTime", "laplacian_variance", "Exposure time versus sharpness"),
        ("BrightnessValue", "skin_linear_luminance_median", "EXIF brightness versus skin luminance"),
    ]
    fig, axes = plt.subplots(1, 3, figsize=(10.0, 3.2))
    for ax, (x_name, y_name, title) in zip(axes, pairs):
        for camera_id, group in formal.groupby("camera_id"):
            color = PALETTE.get(str(camera_id), "#777777")
            x = group[x_name].to_numpy(dtype=float)
            y = group[y_name].to_numpy(dtype=float)
            ax.scatter(x, y, s=11, alpha=0.5, color=color, label=str(camera_id))
            if np.unique(x).size >= 2:
                slope, intercept = np.polyfit(x, y, 1)
                grid = np.linspace(x.min(), x.max(), 100)
                ax.plot(grid, slope * grid + intercept, color=color, lw=1.3)
        ax.set_xlabel(x_name)
        ax.set_ylabel(y_name)
        ax.set_title(title, fontsize=8, fontweight="bold")
        ax.grid(alpha=0.18)
    axes[0].legend(fontsize=6)
    fig.tight_layout()
    save_figure(fig, figures_dir / "exif_vs_aligned_skin_image_quality")


def plot_exif_only_models(
    predictions: pd.DataFrame,
    aggregate_metrics: pd.DataFrame,
    confusion: pd.DataFrame,
    figures_dir: Path,
) -> tuple[str, str, float]:
    ranking = aggregate_metrics.loc[aggregate_metrics["metric"] == "macro_auc"].sort_values(
        "oof_value", ascending=False
    )
    best = ranking.iloc[0]
    model_type = str(best["model_type"])
    feature_group = str(best["feature_group"])
    best_auc = float(best["oof_value"])
    subset = predictions.loc[
        (predictions["model_type"] == model_type) & (predictions["feature_group"] == feature_group)
    ]
    y = subset["label_3class"].to_numpy(dtype=int)
    probabilities = subset[["prob_normal", "prob_mild", "prob_severe"]].to_numpy(dtype=float)

    fig, ax = plt.subplots(figsize=(5.0, 4.2))
    for label, name in CLASS_NAMES.items():
        fpr, tpr, _ = roc_curve((y == label).astype(int), probabilities[:, label])
        auc_value = roc_auc_score((y == label).astype(int), probabilities[:, label])
        ax.plot(fpr, tpr, lw=1.6, color=PALETTE[name], label=f"{name}: AUC={auc_value:.3f}")
    ax.plot([0, 1], [0, 1], ls="--", color="#777777", lw=1)
    ax.set_xlabel("False-positive rate")
    ax.set_ylabel("True-positive rate")
    ax.set_title(f"EXIF-only OOF ROC\n{model_type} / {feature_group}", fontweight="bold")
    ax.legend(fontsize=6.8, loc="lower right")
    ax.set_aspect("equal", adjustable="box")
    ax.grid(alpha=0.16)
    fig.tight_layout()
    save_figure(fig, figures_dir / "exif_only_oof_roc")

    matrix = confusion_matrix(y, probabilities.argmax(axis=1), labels=[0, 1, 2])
    fig, ax = plt.subplots(figsize=(4.4, 3.8))
    image = ax.imshow(matrix, cmap=ListedColormap(["#F3F6FA", "#BCD3E6", "#4C78A8"]))
    maximum = matrix.max()
    for row in range(3):
        for column in range(3):
            color = "white" if matrix[row, column] > maximum * 0.55 else "#222222"
            ax.text(column, row, str(matrix[row, column]), ha="center", va="center", color=color, fontweight="bold")
    labels = [CLASS_NAMES[index] for index in range(3)]
    ax.set_xticks(range(3), labels)
    ax.set_yticks(range(3), labels)
    ax.set_xlabel("Predicted")
    ax.set_ylabel("True")
    ax.set_title(f"EXIF-only OOF confusion matrix\n{model_type} / {feature_group}", fontweight="bold")
    fig.colorbar(image, ax=ax, fraction=0.046, pad=0.04)
    fig.tight_layout()
    save_figure(fig, figures_dir / "exif_only_oof_confusion_matrix")
    return model_type, feature_group, best_auc


def build_known_high_iso_table(
    full: pd.DataFrame,
    formal: pd.DataFrame,
    outliers: pd.DataFrame,
    image_quality: pd.DataFrame,
) -> pd.DataFrame:
    known_ids = outliers.loc[
        (outliers["variable"] == "ISOSpeedRatings")
        & outliers["known_first_stage_iso_outlier_abs_z_gt_5"],
        "ID",
    ].astype(str).tolist()
    if len(known_ids) != 14:
        raise RuntimeError(f"Expected 14 known high-ISO outliers, found {len(known_ids)}: {known_ids}")
    table = full.loc[full["ID"].isin(known_ids)].copy()
    formal_lookup = formal.set_index("ID")["fold"].to_dict()
    table["fold_full"] = table["fold"]
    table["fold_formal"] = table["ID"].map(formal_lookup)
    table["fold_for_current_formal_if_available"] = table["fold_formal"].fillna(table["fold_full"])
    if "high_frequency_noise_mad" not in table.columns:
        table = table.merge(image_quality, on="ID", how="left", validate="one_to_one")
    table["noise_robust_z_within_camera"] = np.nan
    table["blur_robust_z_within_camera"] = np.nan
    full_quality = full.copy()
    if "high_frequency_noise_mad" not in full_quality.columns:
        full_quality = full_quality.merge(image_quality, on="ID", how="left", validate="one_to_one")
    for camera_id, group in full_quality.groupby("camera_id"):
        noise_z = robust_z(group["high_frequency_noise_mad"])
        blur_z = robust_z(group["laplacian_variance"])
        noise_map = dict(zip(group["ID"], noise_z))
        blur_map = dict(zip(group["ID"], blur_z))
        mask = table["camera_id"] == camera_id
        table.loc[mask, "noise_robust_z_within_camera"] = table.loc[mask, "ID"].map(noise_map)
        table.loc[mask, "blur_robust_z_within_camera"] = table.loc[mask, "ID"].map(blur_map)
    table["obvious_overexposure"] = table["overexposed_pixel_ratio"] > 0.01
    table["obvious_underexposure"] = table["underexposed_pixel_ratio"] > 0.05
    table["high_noise_within_camera"] = table["noise_robust_z_within_camera"] > 3.5
    table["obvious_blur_within_camera"] = table["blur_robust_z_within_camera"] < -3.5
    residual_z_map: dict[str, float] = {}
    for _, group in full.groupby("camera_id"):
        residual_z_map.update(dict(zip(group["ID"], robust_z(group["brightness_residual"]))))
    table["brightness_residual_robust_z_within_camera"] = table["ID"].map(residual_z_map)
    table["metadata_inconsistency"] = (
        table["shutter_time_apex_error"].abs() > 1.5
    ) | (table["brightness_residual_robust_z_within_camera"].abs() > 3.5)
    columns = [
        "ID",
        "patient_group_id",
        "camera_id",
        "NYHA",
        "label_3class_name",
        "SEX",
        "fold_full",
        "fold_formal",
        "fold_for_current_formal_if_available",
        *CORE_FIELDS,
        "skin_linear_luminance_median",
        "overexposed_pixel_ratio",
        "underexposed_pixel_ratio",
        "high_frequency_noise_mad",
        "laplacian_variance",
        "noise_robust_z_within_camera",
        "blur_robust_z_within_camera",
        "brightness_residual",
        "shutter_time_apex_error",
        "obvious_overexposure",
        "obvious_underexposure",
        "high_noise_within_camera",
        "obvious_blur_within_camera",
        "metadata_inconsistency",
    ]
    return table[columns].sort_values(["camera_id", "ISOSpeedRatings", "ID"], ascending=[True, False, True])


def build_decision_table(
    full: pd.DataFrame,
    formal: pd.DataFrame,
    by_device: pd.DataFrame,
    apex: pd.DataFrame,
    image_corr: pd.DataFrame,
    confounding: pd.DataFrame,
    device_identity: pd.DataFrame,
) -> pd.DataFrame:
    roles = {
        "ExposureTime": "derived_only",
        "FNumber": "derived_only",
        "ISOSpeedRatings": "derived_only",
        "BrightnessValue": "derived_only",
        "relative_optical_exposure": "core_continuous_condition",
        "log2_iso_gain": "core_continuous_condition",
        "combined_exposure_gain": "quality_control_only",
        "device_centered_brightness": "core_continuous_condition",
        "brightness_residual": "quality_control_only",
    }
    rows = []
    formal_raw_tests = confounding.loc[
        (confounding["cohort"] == "formal_global_meanbg_500")
        & (confounding["outcome"] == "NYHA_3class")
        & (confounding["comparison_context"] == "raw_overall")
        & (confounding["test"] == "Kruskal-Wallis")
    ].set_index("variable")
    identity = device_identity.loc[device_identity["analysis"] == "single_field"].set_index("feature")
    apex_overall = apex.loc[
        (apex["cohort"] == "formal_global_meanbg_500")
        & (apex["analysis_type"] == "fit")
        & (apex["scope"] == "overall")
    ].iloc[0]

    for variable, role in roles.items():
        valid_rate = float(np.isfinite(full[variable].to_numpy(dtype=float)).mean())
        device_stats = by_device.loc[
            (by_device["cohort"] == "formal_global_meanbg_500")
            & (by_device["level"] == "image")
            & (by_device["variable"] == variable)
        ]
        variation_parts = []
        for item in device_stats.itertuples(index=False):
            variation_parts.append(f"{item.camera_id}: unique={int(item.unique_n)}, std={item.std:.4g}, IQR={item.iqr:.4g}")
        identity_row = identity.loc[variable] if variable in identity.index else None
        if identity_row is not None:
            device_dependency = (
                f"single-field camera OOF AUC={identity_row.oof_roc_auc:.3f}; "
                f"MI={identity_row.mutual_information:.3f}"
            )
        else:
            device_dependency = "not estimated"
        if variable in formal_raw_tests.index:
            test_row = formal_raw_tests.loc[variable]
            nyha = (
                f"raw KW epsilon2={test_row.effect_size:.3g}, FDR p={test_row.p_fdr_bh:.3g}, "
                f"minimum overlap={test_row.distribution_overlap_min:.3g}"
            )
        else:
            nyha = "not estimated"

        if variable in CORE_FIELDS:
            relevant = image_corr.loc[
                (image_corr["scope"] == "overall") & (image_corr["exif_field"] == variable)
            ]
            if relevant.empty:
                image_relation = "not separately tested"
            else:
                strongest = relevant.iloc[relevant["spearman_rho"].abs().argmax()]
                image_relation = (
                    f"strongest |rho|: {strongest.image_metric}, rho={strongest.spearman_rho:.3f}"
                )
        else:
            image_relation = "component-derived; no independent image-metric test"

        if variable == "brightness_residual":
            apex_text = "direct APEX consistency residual; quality-control signal"
        elif variable in {"BrightnessValue", "relative_optical_exposure", "log2_iso_gain", "combined_exposure_gain"}:
            apex_text = (
                f"overall recorded-vs-predicted slope={apex_overall.slope:.3f}, R2={apex_overall.r_squared:.3f}; "
                "device-specific fits required"
            )
        elif variable in {"ExposureTime", "FNumber", "ISOSpeedRatings"}:
            apex_text = "valid component of APEX construction; cross-device residual checked separately"
        else:
            apex_text = "not applicable"

        if variable == "ExposureTime":
            conclusion = "保留物理信息，但V1中通过relative_optical_exposure表达，避免与FNumber重复。"
        elif variable == "FNumber":
            conclusion = "设备内变异极低且设备依赖强；只作为relative_optical_exposure组成，不单独学习。"
        elif variable == "ISOSpeedRatings":
            conclusion = "采用log2_iso_gain表达；高ISO保留并作为噪声/质量敏感性标记。"
        elif variable == "BrightnessValue":
            conclusion = "原值不宜跨设备直接使用；V1仅使用训练折设备内稳健中心化值。"
        elif variable == "relative_optical_exposure":
            conclusion = "V1主要光学进光连续条件。"
        elif variable == "log2_iso_gain":
            conclusion = "V1传感器增益连续条件。"
        elif variable == "device_centered_brightness":
            conclusion = "V1亮度条件；每折仅用训练数据的设备中位数/IQR计算。"
        elif variable == "combined_exposure_gain":
            conclusion = "与relative_optical_exposure及ISO确定性冗余，不同时进入V1。"
        else:
            conclusion = "用于APEX/元数据一致性质量控制，不进入表型或标签分类分支。"
        rows.append(
            {
                "字段": variable,
                "完整性": f"{valid_rate:.1%} ({int(valid_rate*len(full))}/{len(full)})",
                "设备内变异": "; ".join(variation_parts),
                "设备依赖": device_dependency,
                "APEX一致性": apex_text,
                "图像表现相关性": image_relation,
                "NYHA混杂": nyha,
                "推荐角色": role,
                "结论依据": conclusion,
            }
        )
    return pd.DataFrame(rows)


def git_commit() -> str:
    try:
        completed = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            cwd=PROJECT_ROOT,
            check=True,
            capture_output=True,
            text=True,
        )
        return completed.stdout.strip() or "unavailable"
    except Exception:
        return "unavailable"


def scalar_metric(
    aggregate: pd.DataFrame,
    model_type: str,
    feature_group: str,
    metric: str,
) -> pd.Series:
    rows = aggregate.loc[
        (aggregate["model_type"] == model_type)
        & (aggregate["feature_group"] == feature_group)
        & (aggregate["metric"] == metric)
    ]
    if len(rows) != 1:
        raise ValueError(f"Expected one metric row for {model_type}/{feature_group}/{metric}, got {len(rows)}")
    return rows.iloc[0]


def write_report(
    path: Path,
    alignment_checks: pd.DataFrame,
    cohorts: dict[str, pd.DataFrame],
    summary: pd.DataFrame,
    apex: pd.DataFrame,
    device_identity: pd.DataFrame,
    auxiliary: pd.DataFrame,
    confounding: pd.DataFrame,
    image_corr: pd.DataFrame,
    known_high_iso: pd.DataFrame,
    aggregate: pd.DataFrame,
    decision: pd.DataFrame,
    best_model_type: str,
    best_feature_group: str,
    best_auc: float,
    image_quality_failures: list[dict[str, Any]],
    output_dir: Path,
) -> dict[str, Any]:
    full = cohorts["all_exif_522"]
    formal = cohorts["formal_global_meanbg_500"]
    full_groups = full["patient_group_id"].nunique()
    formal_groups = formal["patient_group_id"].nunique()
    full_patient = patient_group_frame(full, ANALYSIS_FIELDS)
    formal_patient = patient_group_frame(formal, ANALYSIS_FIELDS)
    full_label_ambiguous = int((full_patient["label_3class_unique_n"] > 1).sum())
    formal_label_ambiguous = int((formal_patient["label_3class_unique_n"] > 1).sum())
    full_camera_ambiguous = int((full_patient["camera_id_unique_n"] > 1).sum())
    formal_camera_ambiguous = int((formal_patient["camera_id_unique_n"] > 1).sum())

    core_summary = summary.loc[
        (summary["cohort"] == "all_exif_522")
        & (summary["level"] == "image")
        & summary["variable"].isin(CORE_FIELDS),
        ["variable", "n", "missing_n", "invalid_n", "min", "p5", "median", "p95", "max", "mean", "std", "iqr", "mad"],
    ]
    camera_counts = formal["camera_id"].value_counts().rename_axis("camera_id").reset_index(name="n")
    fnumber_device = (
        formal.groupby("camera_id")["FNumber"]
        .agg(n="size", unique_n="nunique", min="min", max="max", std="std")
        .reset_index()
    )
    apex_fit = apex.loc[
        (apex["cohort"] == "formal_global_meanbg_500")
        & (apex["analysis_type"] == "fit"),
        ["scope", "camera_id", "n", "pearson_r", "spearman_rho", "slope", "intercept", "r_squared", "residual_median", "residual_mad"],
    ]
    device_top = device_identity.loc[
        device_identity["analysis"] == "single_field",
        ["feature", "hedges_g_b_minus_a", "mutual_information", "oof_roc_auc", "oof_balanced_accuracy"],
    ].sort_values("oof_roc_auc", ascending=False)
    combined_device = device_identity.loc[
        device_identity["analysis"] == "combined_core_and_derived_logistic"
    ].iloc[0]

    nyha_raw = confounding.loc[
        (confounding["cohort"] == "formal_global_meanbg_500")
        & (confounding["outcome"] == "NYHA_3class")
        & (confounding["comparison_context"] == "raw_overall")
        & (confounding["test"] == "Kruskal-Wallis"),
        ["variable", "effect_size", "distribution_overlap_min", "p_value", "p_fdr_bh"],
    ].sort_values("p_fdr_bh")
    nyha_within = confounding.loc[
        (confounding["cohort"] == "formal_global_meanbg_500")
        & (confounding["outcome"] == "NYHA_3class")
        & (confounding["comparison_context"] == "within_camera")
        & (confounding["test"] == "Kruskal-Wallis"),
        ["camera_id", "variable", "effect_size", "distribution_overlap_min", "p_value", "p_fdr_bh"],
    ].sort_values("p_fdr_bh")
    image_corr_view = image_corr.loc[
        image_corr["scope"] == "overall",
        ["exif_field", "image_metric", "n", "spearman_rho", "spearman_p"],
    ].copy()
    image_corr_view["abs_rho"] = image_corr_view["spearman_rho"].abs()
    image_corr_view = image_corr_view.sort_values("abs_rho", ascending=False).drop(columns="abs_rho")

    camera_label_table = pd.crosstab(formal["camera_id"], formal["label_3class_name"])
    camera_label_v, camera_label_p = cramers_v(camera_label_table)
    e3_logistic = scalar_metric(aggregate, "multinomial_logistic", "E3_device_centered_physical", "macro_auc")
    e4_logistic = scalar_metric(aggregate, "multinomial_logistic", "E4_camera_only", "macro_auc")
    e5_logistic = scalar_metric(aggregate, "multinomial_logistic", "E5_E3_plus_camera", "macro_auc")
    best_row = scalar_metric(aggregate, best_model_type, best_feature_group, "macro_auc")

    if best_auc >= 0.75 or float(e4_logistic["oof_value"]) >= 0.70:
        shortcut_level = "明显/严重"
        shortcut_answer = "存在足以否定将raw EXIF直接接入NYHA分类器的明显标签捷径。"
        optical_support = "仅支持把EXIF保留在受约束的物理校正/renderer路线中，不支持作为标签预测分支。"
    elif best_auc >= 0.65 or float(e4_logistic["oof_value"]) >= 0.62:
        shortcut_level = "中等"
        shortcut_answer = "存在需要严肃控制的设备/采集捷径，但尚不足以否定受约束的物理校正路线。"
        optical_support = "支持继续条件光学反演，但必须禁止camera_id/raw EXIF进入NYHA分类分支并做设备敏感性分析。"
    else:
        shortcut_level = "弱"
        shortcut_answer = "未见足以否定物理校正路线的严重标签捷径。"
        optical_support = "支持继续受约束的EXIF条件光学反演；仍需设备分层和外部验证。"

    significant_raw = nyha_raw.loc[nyha_raw["p_fdr_bh"] < 0.05, "variable"].tolist()
    significant_within = nyha_within.loc[nyha_within["p_fdr_bh"] < 0.05, ["camera_id", "variable"]]
    raw_only_mixed = sorted(
        set(significant_raw)
        - set(significant_within["variable"].tolist())
    )

    lines = [
        "# 四个核心EXIF参数二阶段可用性审计",
        "",
        "> 本报告是数据审计与采集捷径诊断，不代表EXIF具有临床预测能力。没有修改原始图像、标签、固定split、历史实验或正式训练代码，也没有启动人脸图像深度学习训练。",
        "",
        "## 结论摘要",
        "",
        f"- 全EXIF队列：{len(full)}张、{full['ID'].nunique()}个唯一ID、{full_groups}个patient_group_id。",
        f"- 当前正式Global/meanbg普通五折队列：{len(formal)}张、{formal['ID'].nunique()}个唯一ID、{formal_groups}个patient_group_id；依据主配置实际指向的splits_500，而非事后筛选的S2 425队列。",
        f"- 四个核心字段在522张中均为100%非空、100%基础合法；两个camera_id在正式队列中的计数见下表。",
        f"- EXIF-only最佳诊断OOF Macro-AUC={best_auc:.3f}（{best_model_type} / {best_feature_group}；patient_group bootstrap 95% CI {best_row.bootstrap_ci_low_95:.3f}–{best_row.bootstrap_ci_high_95:.3f}）。",
        f"- 设备/NYHA采集捷径等级：{shortcut_level}。{shortcut_answer}",
        f"- 最终V1连续EXIF向量：relative_optical_exposure + log2_iso_gain + device_centered_brightness（设备内中位数/IQR只从每折训练部分拟合）。",
        f"- camera_id数据流：只允许作为前向renderer条件及亮度校准索引，不进入表型/NYHA分类编码器。{optical_support}",
        "",
        "## 1. 队列与数据对齐",
        "",
        markdown_table(alignment_checks),
        "",
        f"同患者跨fold：全522队列与正式500队列均未发现。patient_group级连续参数统一取组内中位数。由于NYHA可随同一患者不同照片变化，全队列有{full_label_ambiguous}组、正式队列有{formal_label_ambiguous}组无法安全赋予单一NYHA类别；这些组保留在图像级分析和固定fold OOF中，但不静默压成单一patient_group级NYHA。设备不一致的多图组分别为全队列{full_camera_ambiguous}组、正式队列{formal_camera_ambiguous}组，设备分层patient_group统计不使用这些歧义组。",
        "",
        "正式队列camera_id计数：",
        "",
        markdown_table(camera_counts),
        "",
        "## 2. 四字段基础统计",
        "",
        markdown_table(core_summary),
        "",
        "ExposureTime、FNumber和ISOSpeedRatings同时提供原始尺度与log2尺度统计；完整统计文件还包括图像级、patient_group中位数级、分设备、分NYHA、分SEX及camera_id×NYHA结果。离群值仅标记，没有删除。",
        "",
        "FNumber设备内变异：",
        "",
        markdown_table(fnumber_device),
        "",
        "判断：FNumber在设备内几乎不变，主要编码镜头/设备；仍有明确光学意义，因此只作为relative_optical_exposure的组成，不作为独立学习特征。",
        "",
        "## 3. 派生物理量与APEX一致性",
        "",
        markdown_table(apex_fit),
        "",
        "APEX偏差只用于元数据一致性和设备差异审计。手机ISP、自动曝光和厂商实现会造成截距/斜率差异，不能把残差自动当作源数据错误。BrightnessValue跨设备直接合并会引入系统偏移，V1只使用训练折内按camera_id中位数/IQR稳健中心化后的值。",
        "",
        "## 4. 设备身份编码",
        "",
        markdown_table(device_top.head(12)),
        "",
        f"四字段及派生量联合预测camera_id的固定fold OOF ROC-AUC={combined_device.oof_roc_auc:.3f}、Balanced Accuracy={combined_device.oof_balanced_accuracy:.3f}、Macro-F1={combined_device.oof_macro_f1:.3f}。这说明变量中的设备信息不可忽略；该结果只用于判断采集依赖。",
        "",
        "MeteringMode与Flash未进入任何核心或标签模型。辅助审计显示它们与camera_id的关联及对核心字段方差的解释如下：",
        "",
        markdown_table(auxiliary),
        "",
        "## 5. NYHA与SEX混杂",
        "",
        f"正式队列camera_id×NYHA列联的Cramér's V={camera_label_v:.3f}，卡方p={camera_label_p:.3g}。总体显著但设备内不显著的候选字段：{', '.join(raw_only_mixed) if raw_only_mixed else '无'}。这些候选只能解释为可能的设备混杂，不能解释为心功能相关拍摄差异。",
        "",
        "总体NYHA检验（按FDR排序）：",
        "",
        markdown_table(nyha_raw),
        "",
        "设备内NYHA检验（按FDR排序，前20行）：",
        "",
        markdown_table(nyha_within, max_rows=20),
        "",
        "所有p值均同时报告效应量与经验分布重叠。p<0.05不被自动解释为临床或物理意义；三分类有序Spearman仅为探索性统计。完整SEX、设备内及设备调整后结果见core_parameter_confounding_tests.csv。",
        "",
        "## 6. 与实际图像表现的关系",
        "",
        "使用现有224×224颜色保持aligned_rgb及既有CelebAMask-HQ解析标签中的skin类计算；未重新训练或生成分割网络。颜色/亮度统计没有使用meanbg图、black-background成图或ImageNet标准化张量。",
        "",
        f"图像质量计算失败：{len(image_quality_failures)}例。主要Spearman相关如下：",
        "",
        markdown_table(image_corr_view),
        "",
        "质量指标定义：skin区域中位RGB；sRGB反伽马后的相对线性亮度；由线性Y换算的Lab L*；线性亮度≥0.98为过曝、≤0.02为欠曝；任一通道≤5或≥250为饱和/裁剪；Laplacian方差为清晰度代理；高斯平滑残差MAD为简化高频噪声代理。它们是相对质量指标，不是传感器标定真值。",
        "",
        "已知14张设备内高ISO离群图片均保留。逐例过曝、欠曝、噪声、模糊和元数据不一致标记见known_high_iso_outliers.csv；判定使用固定阈值和设备内稳健z，不把统计标记等同于源图错误。",
        "",
        markdown_table(known_high_iso[["ID", "patient_group_id", "camera_id", "label_3class_name", "SEX", "fold_for_current_formal_if_available", "ISOSpeedRatings", "obvious_overexposure", "obvious_underexposure", "high_noise_within_camera", "obvious_blur_within_camera", "metadata_inconsistency"]]),
        "",
        "## 7. EXIF-only固定五折标签捷径诊断",
        "",
        f"Logistic E3 OOF Macro-AUC={e3_logistic.oof_value:.3f}（95% CI {e3_logistic.bootstrap_ci_low_95:.3f}–{e3_logistic.bootstrap_ci_high_95:.3f}）；camera-only E4={e4_logistic.oof_value:.3f}；E3+camera E5={e5_logistic.oof_value:.3f}。所有标准化、填补、camera编码和设备内亮度中位数/IQR均只在各折训练部分拟合。",
        "",
        "OOF文件为每个固定特征组和两种预设模型保留完整三分类概率。每个ID在每个模型/特征组中恰好出现一次、概率有限且行和为1；同患者不跨fold。模型和特征未根据结果调参。",
        "",
        "## 8. 共线性与字段压缩",
        "",
        "ExposureTime、FNumber、ISO及派生曝光量包含确定性或近确定性关系，VIF出现无穷大属于公式冗余的预期结果，不机械按阈值删字段。V1选择relative_optical_exposure + log2_iso_gain + device_centered_brightness，避免同时输入原始四字段、EV100及combined_exposure_gain。",
        "",
        "## 9. 最终逐字段决策表",
        "",
        markdown_table(decision),
        "",
        "## 10. 对八个问题的明确回答",
        "",
        "1. ExposureTime不作为独立原始条件进入V1，只通过relative_optical_exposure使用。",
        "2. FNumber设备内变异极低，本质上接近设备/镜头常量；仅保留其物理公式作用。",
        "3. ISOSpeedRatings存在真实设备内变化；是否反映噪声以图像高频噪声相关为证据，但手机ISP降噪会削弱对应关系，因此采用log2_iso_gain并保留高ISO质量标记。",
        "4. BrightnessValue不能跨设备直接使用，必须在每折训练数据内按设备稳健中心化/缩放。",
        "5. camera_id不进入光学反演/表型/分类编码器，只进入前向renderer；它可以作为设备内亮度校准索引，但不作为可学习标签捷径输入。",
        "6. V1连续EXIF向量为[relative_optical_exposure, log2_iso_gain, device_centered_brightness]。",
        "7. combined_exposure_gain与brightness_residual只用于质量控制/敏感性；原始ExposureTime、FNumber、ISO、BrightnessValue只作为派生来源或审计输出。",
        f"8. {shortcut_answer} {optical_support}",
        "",
        "## 11. 文件说明",
        "",
        "所有CSV均为UTF-8-SIG。figures目录同时保存PNG和可编辑文字SVG。报告目录：" + str(output_dir.resolve()),
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return {
        "shortcut_level": shortcut_level,
        "shortcut_answer": shortcut_answer,
        "optical_support": optical_support,
        "camera_label_cramers_v": camera_label_v,
        "camera_label_chi_square_p": camera_label_p,
        "best_model_type": best_model_type,
        "best_feature_group": best_feature_group,
        "best_macro_auc": best_auc,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--metadata-workbook",
        type=Path,
        default=PROJECT_ROOT / "data/raw/EXIF/Image_Metadata_All.xlsx",
    )
    parser.add_argument(
        "--exif-checklist",
        type=Path,
        default=PROJECT_ROOT / "data/raw/EXIF/EXIF_Inform.csv",
    )
    parser.add_argument(
        "--full-split",
        type=Path,
        default=PROJECT_ROOT / "data/processed/splits/nyha_3class_sex_stratified_group_5fold.csv",
    )
    parser.add_argument(
        "--formal-split",
        type=Path,
        default=PROJECT_ROOT / "data/processed/splits_500/nyha_3class_sex_stratified_group_5fold.csv",
    )
    parser.add_argument(
        "--formal-config",
        type=Path,
        default=PROJECT_ROOT
        / "config/train/preprocess_ablation_resnet18/nyha_3class_resnet18_preproc_hybrid_imagenet_meanbg.yaml",
    )
    parser.add_argument(
        "--formal-meanbg-dir",
        type=Path,
        default=PROJECT_ROOT / "data/processed/global_face/preprocess_ablation/hybrid_imagenet_meanbg/images",
    )
    parser.add_argument(
        "--aligned-rgb-dir",
        type=Path,
        default=PROJECT_ROOT
        / "data/processed/global_face/global_face_parsing_hybrid_foreheadrepair_blackbg_224_png_strict_intermediates/aligned_rgb",
    )
    parser.add_argument(
        "--parsing-label-dir",
        type=Path,
        default=PROJECT_ROOT
        / "data/processed/global_face/global_face_parsing_hybrid_foreheadrepair_blackbg_224_png_strict_intermediates/parsing_label",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=PROJECT_ROOT / "reports/exif_core_parameter_second_stage_audit",
    )
    parser.add_argument("--overwrite-output", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    configure_plotting()
    required_files = [
        args.metadata_workbook,
        args.exif_checklist,
        args.full_split,
        args.formal_split,
        args.formal_config,
    ]
    missing = [str(path) for path in required_files if not path.is_file()]
    required_dirs = [args.formal_meanbg_dir, args.aligned_rgb_dir, args.parsing_label_dir]
    missing.extend(str(path) for path in required_dirs if not path.is_dir())
    if missing:
        raise FileNotFoundError(f"Required inputs missing: {missing}")

    output_dir = args.output_dir.resolve()
    if output_dir.exists() and any(output_dir.iterdir()) and not args.overwrite_output:
        raise FileExistsError(
            f"Output directory already contains files: {output_dir}. Use --overwrite-output only for this audit package."
        )
    figures_dir = output_dir / "figures"
    figures_dir.mkdir(parents=True, exist_ok=True)

    exif, exif_conflicts = load_exif_values(args.metadata_workbook)
    full_split = read_split(args.full_split, "all_exif_522")
    formal_split = read_split(args.formal_split, "formal_global_meanbg_500")
    alignment, alignment_checks = check_cohort_alignment(exif, full_split, formal_split)

    formal_config_text = args.formal_config.read_text(encoding="utf-8")
    config_matches_split = "data/processed/splits_500" in formal_config_text.replace("\\", "/")
    config_matches_meanbg = "hybrid_imagenet_meanbg/images" in formal_config_text.replace("\\", "/")
    formal_images_missing = [
        image_id for image_id in formal_split["ID"].astype(str) if not (args.formal_meanbg_dir / f"{image_id}.png").is_file()
    ]
    aligned_missing = [
        image_id for image_id in full_split["ID"].astype(str) if not (args.aligned_rgb_dir / f"{image_id}.png").is_file()
    ]
    label_missing = [
        image_id for image_id in full_split["ID"].astype(str) if not (args.parsing_label_dir / f"{image_id}.png").is_file()
    ]
    core_numeric = exif[CORE_FIELDS].apply(pd.to_numeric, errors="coerce")
    core_valid = np.isfinite(core_numeric.to_numpy()).all(axis=1)
    core_valid &= (core_numeric[["ExposureTime", "FNumber", "ISOSpeedRatings"]] > 0).all(axis=1).to_numpy()
    extra_checks = [
        {
            "check": "formal_config_points_to_splits_500",
            "status": "PASS" if config_matches_split else "FAIL",
            "detail": str(args.formal_config.resolve()),
        },
        {
            "check": "formal_config_points_to_meanbg",
            "status": "PASS" if config_matches_meanbg else "FAIL",
            "detail": str(args.formal_meanbg_dir.resolve()),
        },
        {
            "check": "formal_meanbg_images_available",
            "status": "PASS" if not formal_images_missing else "FAIL",
            "detail": f"missing_n={len(formal_images_missing)}; IDs={formal_images_missing}",
        },
        {
            "check": "aligned_rgb_available_for_full_queue",
            "status": "PASS" if not aligned_missing else "FAIL",
            "detail": f"missing_n={len(aligned_missing)}; IDs={aligned_missing}",
        },
        {
            "check": "parsing_label_available_for_full_queue",
            "status": "PASS" if not label_missing else "FAIL",
            "detail": f"missing_n={len(label_missing)}; IDs={label_missing}",
        },
        {
            "check": "core_exif_values_complete_and_positive",
            "status": "PASS" if bool(core_valid.all()) else "FAIL",
            "detail": f"valid_n={int(core_valid.sum())}/{len(exif)}; invalid_IDs={exif.loc[~core_valid, 'ID'].tolist()}",
        },
        {
            "check": "no_conflicting_duplicate_core_or_aux_tags",
            "status": "PASS" if not exif_conflicts else "FAIL",
            "detail": f"conflicts={exif_conflicts}",
        },
    ]
    alignment_checks = pd.concat([alignment_checks, pd.DataFrame(extra_checks)], ignore_index=True)
    save_csv(alignment, output_dir / "cohort_alignment_audit.csv")
    save_csv(alignment_checks, output_dir / "cohort_alignment_checks.csv")
    failed_checks = alignment_checks.loc[alignment_checks["status"] != "PASS"]
    if not failed_checks.empty:
        raise RuntimeError("Cohort/input alignment failed:\n" + failed_checks.to_string(index=False))

    cohorts = assemble_cohorts(exif, full_split, formal_split)
    full = cohorts["all_exif_522"].copy()
    formal = cohorts["formal_global_meanbg_500"].copy()
    camera_counts_full = full["camera_id"].value_counts().to_dict()
    if len(camera_counts_full) != 2:
        raise RuntimeError(f"Expected two camera_id values, got {camera_counts_full}")

    print("[stage] computing aligned-skin image quality metrics", flush=True)
    image_quality, image_quality_failures = compute_image_quality(full, args.aligned_rgb_dir, args.parsing_label_dir)
    save_csv(pd.DataFrame(image_quality_failures), output_dir / "image_quality_failures.csv")
    if image_quality_failures:
        raise RuntimeError(
            f"Image quality analysis stopped: {len(image_quality_failures)} aligned image/mask failures; see image_quality_failures.csv"
        )
    full = full.merge(image_quality, on="ID", how="left", validate="one_to_one")
    formal = formal.merge(image_quality, on="ID", how="left", validate="one_to_one")
    cohorts = {"all_exif_522": full, "formal_global_meanbg_500": formal}

    print("[stage] descriptive, APEX, device, confounding and collinearity analyses", flush=True)
    summary = summarize_overall(cohorts)
    by_device = summarize_grouped(cohorts, ["camera_id"])
    by_nyha = summarize_grouped(cohorts, ["label_3class", "label_3class_name"])
    by_sex = summarize_grouped(cohorts, ["SEX", "sex_name"])
    by_camera_nyha = summarize_grouped(cohorts, ["camera_id", "label_3class", "label_3class_name"])
    outliers = build_outlier_table(full, set(formal["ID"]))
    apex = apex_consistency(cohorts)
    device_identity = device_identity_analysis(formal)
    auxiliary, auxiliary_crosstab = auxiliary_device_analysis(formal)
    confounding = confounding_tests(cohorts)
    image_corr = image_quality_correlations(formal)
    correlations = correlation_tables(cohorts)
    vif = vif_tables(cohorts)
    known_high_iso = build_known_high_iso_table(full, formal, outliers, image_quality)

    print("[stage] fixed-fold EXIF-only OOF diagnostics and 2000-repeat group bootstrap", flush=True)
    predictions, aggregate, fold_metrics, confusion, model_audit = exif_only_oof(formal.reset_index(drop=True))
    oof_validation = pd.DataFrame(model_audit["validation"])
    if not (
        oof_validation["each_id_once"].all()
        and oof_validation["probabilities_finite"].all()
        and (~oof_validation["same_patient_cross_fold"]).all()
        and (oof_validation["max_probability_sum_abs_error"] < 1e-8).all()
    ):
        raise RuntimeError("OOF validation failed; see exif_only_oof_validation.csv")

    decision = build_decision_table(
        full, formal, by_device, apex, image_corr, confounding, device_identity
    )

    values_output = full.copy()
    formal_fold_lookup = formal.set_index("ID")["fold"].to_dict()
    values_output = values_output.rename(columns={"fold": "fold_full_522"})
    values_output["in_formal_global_meanbg_500"] = values_output["ID"].isin(set(formal["ID"]))
    values_output["fold_formal_global_meanbg_500"] = values_output["ID"].map(formal_fold_lookup)
    value_columns = [
        "ID",
        "patient_group_id",
        "NYHA",
        "label_3class",
        "label_3class_name",
        "SEX",
        "sex_name",
        "fold_full_522",
        "in_formal_global_meanbg_500",
        "fold_formal_global_meanbg_500",
        "Make",
        "Model",
        "camera_id",
        *CORE_FIELDS,
        "MeteringMode",
        "Flash",
        "ShutterSpeedValue",
        *DERIVED_FIELDS,
        "shutter_time_apex_error",
        *[column for column in image_quality.columns if column != "ID"],
    ]
    save_csv(values_output[value_columns], output_dir / "core_parameter_values_and_derived.csv")
    save_csv(summary, output_dir / "core_parameter_summary.csv")
    save_csv(by_device, output_dir / "core_parameter_by_device.csv")
    save_csv(by_nyha, output_dir / "core_parameter_by_nyha.csv")
    save_csv(by_sex, output_dir / "core_parameter_by_sex.csv")
    save_csv(by_camera_nyha, output_dir / "core_parameter_by_camera_nyha.csv")
    save_csv(outliers, output_dir / "core_parameter_outliers.csv")
    save_csv(apex, output_dir / "core_parameter_apex_consistency.csv")
    save_csv(device_identity, output_dir / "core_parameter_device_identity.csv")
    save_csv(auxiliary, output_dir / "auxiliary_exif_device_explanation.csv")
    save_csv(auxiliary_crosstab, output_dir / "auxiliary_exif_device_crosstab.csv")
    save_csv(confounding, output_dir / "core_parameter_confounding_tests.csv")
    save_csv(image_quality, output_dir / "aligned_skin_image_quality_metrics.csv")
    save_csv(image_corr, output_dir / "core_parameter_image_quality_correlations.csv")
    save_csv(known_high_iso, output_dir / "known_high_iso_outliers.csv")
    save_csv(correlations, output_dir / "core_parameter_correlations.csv")
    save_csv(vif, output_dir / "core_parameter_vif.csv")
    save_csv(predictions, output_dir / "exif_only_oof_predictions.csv")
    save_csv(aggregate, output_dir / "exif_only_cv_metrics.csv")
    save_csv(fold_metrics, output_dir / "exif_only_fold_metrics.csv")
    save_csv(confusion, output_dir / "exif_only_confusion_matrices.csv")
    save_csv(oof_validation, output_dir / "exif_only_oof_validation.csv")
    save_csv(decision, output_dir / "core_parameter_decision_table.csv")

    print("[stage] generating figures", flush=True)
    plot_core_distributions(formal, figures_dir)
    plot_core_by_nyha(formal, figures_dir)
    plot_camera_nyha(formal, figures_dir)
    plot_correlation_heatmap(formal, figures_dir)
    plot_apex(formal, figures_dir)
    plot_image_quality_relations(formal, figures_dir)
    best_model_type, best_feature_group, best_auc = plot_exif_only_models(
        predictions, aggregate, confusion, figures_dir
    )

    report_summary = write_report(
        output_dir / "exif_core_parameter_audit_report.md",
        alignment_checks,
        cohorts,
        summary,
        apex,
        device_identity,
        auxiliary,
        confounding,
        image_corr,
        known_high_iso,
        aggregate,
        decision,
        best_model_type,
        best_feature_group,
        best_auc,
        image_quality_failures,
        output_dir,
    )

    input_files = [
        args.metadata_workbook,
        args.exif_checklist,
        args.full_split,
        args.formal_split,
        args.formal_config,
        PROJECT_ROOT / "data/raw/label_raw.csv",
        PROJECT_ROOT / "data/raw/label_raw_nyha2_remove22_sex_balanced_500.csv",
        PROJECT_ROOT / "scripts/5fold/build_nyha_3class_sex_stratified_group_5fold_split.py",
        PROJECT_ROOT / "scripts/analysis/audit_exif_parameter_completeness.py",
    ]
    manifest = {
        "name": "exif_core_parameter_second_stage_audit",
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "overall_status": "COMPLETE",
        "project_root": str(PROJECT_ROOT.resolve()),
        "output_dir": str(output_dir),
        "input_files": {
            path.name: {"path": str(path.resolve()), "sha256": sha256_file(path)}
            for path in input_files
            if path.is_file()
        },
        "input_directories": {
            "formal_meanbg_images": sha256_directory(args.formal_meanbg_dir),
            "aligned_rgb": sha256_directory(args.aligned_rgb_dir),
            "parsing_label": sha256_directory(args.parsing_label_dir),
        },
        "cohorts": {
            "all_exif_522": {
                "definition": "All IDs in Image_Metadata_All.xlsx aligned one-to-one to the fixed 522 master split",
                "image_n": len(full),
                "unique_id_n": full["ID"].nunique(),
                "patient_group_n": full["patient_group_id"].nunique(),
            },
            "formal_global_meanbg_500": {
                "definition": "Current Global/meanbg main config data.split_dir=data/processed/splits_500",
                "image_n": len(formal),
                "unique_id_n": formal["ID"].nunique(),
                "patient_group_n": formal["patient_group_id"].nunique(),
                "split_path": str(args.formal_split.resolve()),
                "split_sha256": sha256_file(args.formal_split),
                "image_root": str(args.formal_meanbg_dir.resolve()),
            },
            "excluded_interpretation": "S2_425 is post-hoc and was not used as the current formal cohort",
        },
        "label_mapping": {
            "NYHA_0": "normal/0",
            "NYHA_1_or_2": "mild/1",
            "NYHA_3_or_4": "severe/2",
            "SEX_0": "female",
            "SEX_1": "male",
        },
        "patient_group_source": "split patient_group_id; image suffix -digits removed by authoritative split builder",
        "fields": {"core": CORE_FIELDS, "auxiliary_audit_only": AUX_FIELDS, "derived": DERIVED_FIELDS},
        "derived_formulas": {
            "log2_exposure_time": "log2(ExposureTime)",
            "log2_iso_gain": "log2(ISOSpeedRatings/100)",
            "aperture_value_from_f": "2*log2(FNumber)",
            "time_value_from_t": "-log2(ExposureTime)",
            "EV100": "aperture_value_from_f + time_value_from_t",
            "relative_optical_exposure": "log2(ExposureTime) - 2*log2(FNumber)",
            "combined_exposure_gain": "relative_optical_exposure + log2_iso_gain",
            "sensitivity_value": "log2(ISOSpeedRatings/3.125)",
            "brightness_apex_pred": "aperture_value_from_f + time_value_from_t - sensitivity_value",
            "brightness_residual": "BrightnessValue - brightness_apex_pred",
            "device_centered_brightness_descriptive": "(BrightnessValue - full-cohort camera median)/camera IQR",
            "device_centered_brightness_cv": "(BrightnessValue - training-fold camera median)/training-fold camera IQR",
        },
        "statistical_methods": {
            "descriptive": "image-level and patient_group median; quantiles, IQR, raw MAD",
            "outliers": "within-camera IQR 1.5 rule and robust z=0.67448975*(x-median)/MAD, abs(z)>3.5; known high ISO abs(z)>5",
            "APEX": "Pearson, Spearman, OLS slope/intercept/R2 and camera interaction",
            "NYHA": "Kruskal-Wallis, epsilon squared, empirical histogram overlap, BH-FDR",
            "SEX": "Mann-Whitney U, rank-biserial effect, empirical histogram overlap, BH-FDR",
            "device_adjustment": "camera-median centered exploratory rank tests",
            "bootstrap": f"patient_group_id cluster bootstrap, repeats={BOOTSTRAP_REPEATS}, seed={SEED}",
            "correlation": "Pearson and Spearman overall and within camera",
            "VIF": "1/(1-R2), exact deterministic dependencies reported as infinity",
        },
        "image_quality": {
            "source": "existing color-preserving aligned_rgb with existing parsing_label class 1 skin mask",
            "not_used": ["raw full-frame without aligned mask", "meanbg", "black-background image", "ImageNet-normalized tensor"],
            "metrics": [
                "median RGB",
                "sRGB-linearized relative luminance",
                "Lab L* from linear Y",
                "over/underexposure ratios",
                "channel clipping ratio",
                "Laplacian variance",
                "Gaussian high-frequency residual MAD",
                "skin pixel ratio",
            ],
        },
        "models": {
            "feature_groups": FEATURE_GROUPS,
            "logistic": {
                "penalty": "l2",
                "C": 1.0,
                "solver": "lbfgs",
                "class_weight": "balanced",
                "random_state": SEED,
                "max_iter": 3000,
            },
            "random_forest": {
                "n_estimators": 300,
                "max_depth": 3,
                "min_samples_leaf": 20,
                "class_weight": "balanced",
                "random_state": SEED,
            },
            "fold_preprocessing_audit": model_audit["fold_preprocessing"],
        },
        "random_seed": SEED,
        "bootstrap_repeats": BOOTSTRAP_REPEATS,
        "software": {
            "python": sys.version,
            "python_executable": sys.executable,
            "platform": platform.platform(),
            "pandas": pd.__version__,
            "numpy": np.__version__,
            "scipy": scipy.__version__,
            "scikit_learn": sklearn.__version__,
            "matplotlib": matplotlib.__version__,
            "statsmodels": statsmodels.__version__,
            "openpyxl": openpyxl.__version__,
            "Pillow": PIL.__version__,
        },
        "git_commit": git_commit(),
        "alignment_checks": alignment_checks.to_dict(orient="records"),
        "oof_validation": model_audit["validation"],
        "report_summary": report_summary,
        "prohibitions_observed": {
            "source_files_modified": False,
            "fixed_splits_regenerated": False,
            "outliers_deleted": False,
            "deep_image_training_started": False,
            "dependencies_installed_or_upgraded": False,
        },
    }
    (output_dir / "run_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, allow_nan=False, default=str) + "\n",
        encoding="utf-8",
    )

    roles = dict(zip(decision["字段"], decision["推荐角色"]))
    best = report_summary
    print("\n=== EXIF CORE SECOND-STAGE AUDIT ===")
    print("Overall status: COMPLETE")
    print(f"All EXIF cohort: images={len(full)}, unique_IDs={full['ID'].nunique()}, patient_groups={full['patient_group_id'].nunique()}")
    print(f"Formal Global/meanbg cohort: images={len(formal)}, unique_IDs={formal['ID'].nunique()}, patient_groups={formal['patient_group_id'].nunique()}")
    for field in CORE_FIELDS:
        values = pd.to_numeric(full[field], errors="coerce")
        complete = values.notna().mean()
        legal = (values.notna() & np.isfinite(values)).mean()
        if field in {"ExposureTime", "FNumber", "ISOSpeedRatings"}:
            legal = (values.notna() & np.isfinite(values) & (values > 0)).mean()
        print(f"{field}: completeness={complete:.1%}, legality={legal:.1%}, role={roles[field]}")
    for camera_id, count in full["camera_id"].value_counts().items():
        print(f"camera_id {camera_id}: n={count}")
    print("Patient group cross-fold leakage: NO")
    for field in ["relative_optical_exposure", "log2_iso_gain", "combined_exposure_gain", "device_centered_brightness", "brightness_residual"]:
        print(f"{field}: role={roles[field]}")
    print("Recommended V1 continuous EXIF vector: relative_optical_exposure, log2_iso_gain, device_centered_brightness")
    print("camera_id flow: renderer-only plus non-learned fold-safe calibration index; not phenotype/classifier encoder")
    print(f"Best EXIF-only diagnostic Macro-AUC: {best_auc:.6f} ({best_model_type}/{best_feature_group})")
    print(f"Device/NYHA shortcut: {best['shortcut_level']} - {best['shortcut_answer']}")
    print(f"Continue EXIF-conditioned optical inversion: {best['optical_support']}")
    print(f"Report: {output_dir / 'exif_core_parameter_audit_report.md'}")
    print(f"Key values CSV: {output_dir / 'core_parameter_values_and_derived.csv'}")
    print(f"OOF predictions CSV: {output_dir / 'exif_only_oof_predictions.csv'}")
    print(f"Decision table CSV: {output_dir / 'core_parameter_decision_table.csv'}")


if __name__ == "__main__":
    main()
