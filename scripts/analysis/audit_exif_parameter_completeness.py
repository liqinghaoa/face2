"""Audit EXIF checklist completeness and plausibility at image and patient level."""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
PLACEHOLDERS = {"", "none", "nan", "n/a", "na", "null", "unknown", "<blank>"}
NUMERIC_PARAMETERS = {
    "Orientation", "ColorSpace", "XResolution", "YResolution", "ResolutionUnit",
    "YCbCrPositioning", "ExifImageWidth", "ExifImageHeight", "ExposureTime",
    "FNumber", "ISOSpeedRatings", "ISOSpeed", "FocalLength",
    "FocalLengthIn35mmFilm", "ExposureBiasValue", "ExposureProgram",
    "ExposureMode", "MeteringMode", "LightSource", "Flash", "WhiteBalance",
    "BrightnessValue", "ShutterSpeedValue", "ApertureValue", "MaxApertureValue",
    "DigitalZoomRatio", "SceneCaptureType", "SensingMethod",
}
CORE_DIRECT = {
    "Make", "Model", "DateTimeOriginal", "Orientation", "ExifImageWidth",
    "ExifImageHeight", "ExposureTime", "FNumber", "FocalLength",
    "ExposureBiasValue", "ExposureProgram", "ExposureMode", "MeteringMode",
    "LightSource", "Flash", "WhiteBalance", "BrightnessValue",
    "ShutterSpeedValue", "SceneCaptureType",
}


def save_csv(frame: pd.DataFrame, path: Path) -> None:
    frame.to_csv(path, index=False, encoding="utf-8-sig")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.startswith("'"):
        text = text[1:].strip()
    return text


def is_blank(value: Any) -> bool:
    return clean_text(value).casefold() in PLACEHOLDERS


def parse_number(value: Any) -> float | None:
    text = clean_text(value).replace(",", "")
    if text.casefold() in PLACEHOLDERS:
        return None
    try:
        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?/[-+]?\d+(?:\.\d+)?", text):
            numerator, denominator = text.split("/", 1)
            denominator_value = float(denominator)
            return float(numerator) / denominator_value if denominator_value != 0 else None
        value_float = float(text)
        return value_float if math.isfinite(value_float) else None
    except ValueError:
        return None


def validate_parameter(parameter: str, value: Any) -> tuple[str, str, float | None]:
    """Return status, reason, and parsed numeric value when applicable."""
    text = clean_text(value)
    if is_blank(value):
        return "blank", "empty_or_placeholder", None
    if parameter in {"Make", "Model", "Software", "ImageDescription"}:
        return "valid", "nonempty_text", None
    if parameter in {"DateTime", "DateTimeOriginal", "DateTimeDigitized"}:
        try:
            parsed = datetime.strptime(text, "%Y:%m:%d %H:%M:%S")
        except ValueError:
            return "invalid", "invalid_EXIF_datetime_format", None
        if parsed < datetime(2000, 1, 1) or parsed > datetime.now() + timedelta(days=2):
            return "warning", "datetime_outside_2000_to_extraction_period", parsed.timestamp()
        return "valid", "valid_EXIF_datetime", parsed.timestamp()
    if parameter in {"OffsetTime", "OffsetTimeOriginal"}:
        match = re.fullmatch(r"([+-])(\d{2}):(\d{2})", text)
        if not match:
            return "invalid", "invalid_timezone_offset", None
        hours, minutes = int(match.group(2)), int(match.group(3))
        return ("valid", "valid_timezone_offset", None) if hours <= 14 and minutes < 60 else ("invalid", "timezone_offset_out_of_range", None)
    if parameter in {"ExifVersion", "FlashPixVersion"}:
        return ("valid", "four_digit_version", None) if re.fullmatch(r"\d{4}", text) else ("invalid", "invalid_version_format", None)
    if parameter in {"SubsecTime", "SubsecTimeOriginal", "SubsecTimeDigitized"}:
        return ("valid", "numeric_subsecond_text", None) if re.fullmatch(r"\d{1,9}", text) else ("invalid", "invalid_subsecond_text", None)

    number = parse_number(value)
    if parameter in NUMERIC_PARAMETERS and number is None:
        return "invalid", "not_parseable_as_number", None
    if parameter not in NUMERIC_PARAMETERS:
        return "valid", "nonempty_unrestricted_text", None

    ranges: dict[str, tuple[float, float]] = {
        "Orientation": (1, 8), "ColorSpace": (1, 65535), "XResolution": (0.01, 100000),
        "YResolution": (0.01, 100000), "ResolutionUnit": (1, 3), "YCbCrPositioning": (1, 2),
        "ExifImageWidth": (1, 100000), "ExifImageHeight": (1, 100000),
        "ExposureTime": (1e-7, 60), "FNumber": (0.5, 64),
        "ISOSpeedRatings": (1, 204800), "ISOSpeed": (1, 204800),
        "FocalLength": (0.1, 1000), "FocalLengthIn35mmFilm": (1, 2000),
        "ExposureBiasValue": (-10, 10), "ExposureProgram": (0, 8),
        "ExposureMode": (0, 2), "MeteringMode": (0, 255), "LightSource": (0, 255),
        "Flash": (0, 255), "WhiteBalance": (0, 1), "BrightnessValue": (-100, 100),
        "ShutterSpeedValue": (-50, 50), "ApertureValue": (0.01, 32),
        "MaxApertureValue": (0.01, 32), "DigitalZoomRatio": (0, 1000),
        "SceneCaptureType": (0, 3), "SensingMethod": (1, 8),
    }
    low, high = ranges[parameter]
    if not low <= number <= high:
        return "invalid", f"outside_plausible_range_{low}_to_{high}", number

    enum_values = {
        "Orientation": set(range(1, 9)), "ColorSpace": {1, 65535}, "ResolutionUnit": {1, 2, 3},
        "YCbCrPositioning": {1, 2}, "ExposureProgram": set(range(0, 9)),
        "ExposureMode": {0, 1, 2}, "MeteringMode": {0, 1, 2, 3, 4, 5, 6, 255},
        "LightSource": set(range(0, 25)) | {255}, "WhiteBalance": {0, 1},
        "SceneCaptureType": {0, 1, 2, 3}, "SensingMethod": set(range(1, 9)),
    }
    if parameter in enum_values and int(number) not in enum_values[parameter]:
        return "invalid", "undefined_enum_code", number

    if parameter == "DigitalZoomRatio" and number > 20:
        return "warning", "implausibly_large_or_device_scaled_zoom_ratio", number
    if (parameter, int(number)) in {("ExposureProgram", 0), ("LightSource", 255), ("SensingMethod", 1)}:
        return "warning", "legal_but_noninformative_enum_code", number
    return "valid", "within_rule", number


def load_workbook_data(path: Path, target_parameters: set[str]) -> tuple[pd.DataFrame, dict[str, dict[str, Any]], dict[tuple[str, str], list[str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    metadata_sheet = wb.worksheets[1]
    iterator = metadata_sheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else f"unnamed_{index}" for index, value in enumerate(next(iterator))]
    metadata = pd.DataFrame(list(iterator), columns=headers)
    metadata["ID"] = metadata["ID"].astype(str).str.strip()

    raw_sheet = wb.worksheets[2]
    iterator = raw_sheet.iter_rows(values_only=True)
    next(iterator)
    values: dict[str, dict[str, Any]] = defaultdict(dict)
    duplicates: dict[tuple[str, str], list[str]] = defaultdict(list)
    for row in iterator:
        patient_id = clean_text(row[0])
        parameter = clean_text(row[4])
        if parameter not in target_parameters:
            continue
        raw_value = row[5]
        key = (patient_id, parameter)
        duplicates[key].append(clean_text(raw_value))
        if parameter not in values[patient_id] or is_blank(values[patient_id][parameter]):
            values[patient_id][parameter] = raw_value
    wb.close()
    conflicting = {key: sorted(set(items)) for key, items in duplicates.items() if len(set(items)) > 1}
    return metadata, values, conflicting


def load_group_mapping(path: Path, image_ids: set[str]) -> tuple[dict[str, str], str]:
    if not path.is_file():
        return {patient_id: patient_id for patient_id in image_ids}, "fallback_to_image_ID"
    frame = pd.read_csv(path, dtype={"ID": "string", "patient_group_id": "string"}, encoding="utf-8-sig")
    mapping = dict(zip(frame["ID"].astype(str), frame["patient_group_id"].astype(str)))
    missing = image_ids - set(mapping)
    for patient_id in missing:
        mapping[patient_id] = patient_id
    return mapping, f"project_patient_group_id; fallback_for_{len(missing)}_unmapped_images"


def add_cross_field_issues(image: pd.Series, parameter_results: dict[str, dict[str, Any]], issues: list[dict[str, Any]]) -> None:
    patient_id = str(image["ID"])
    result = parameter_results[patient_id]
    def number(name: str) -> float | None:
        item = result.get(name)
        return item.get("numeric_value") if item else None
    def issue(parameter: str, issue_type: str, detail: str) -> None:
        issues.append({"ID": patient_id, "patient_group_id": image["patient_group_id"], "parameter": parameter, "issue_type": issue_type, "severity": "review", "raw_value": result.get(parameter, {}).get("raw_value", ""), "detail": detail})

    actual_width, actual_height = parse_number(image.get("宽度(px)")), parse_number(image.get("高度(px)"))
    exif_width, exif_height = number("ExifImageWidth"), number("ExifImageHeight")
    if None not in {actual_width, actual_height, exif_width, exif_height}:
        direct = actual_width == exif_width and actual_height == exif_height
        swapped = actual_width == exif_height and actual_height == exif_width
        if not direct and not swapped:
            issue("ExifImageWidth/Height", "dimension_mismatch", f"actual={actual_width}x{actual_height}, EXIF={exif_width}x{exif_height}")
    exposure, shutter = number("ExposureTime"), number("ShutterSpeedValue")
    if exposure and shutter is not None:
        expected = -math.log2(exposure)
        if abs(expected - shutter) > 1.5:
            issue("ExposureTime/ShutterSpeedValue", "cross_field_inconsistency", f"expected_APEX={expected:.3f}, recorded={shutter:.3f}")
    fnumber, aperture = number("FNumber"), number("ApertureValue")
    if fnumber and aperture is not None:
        expected = 2 * math.log2(fnumber)
        if abs(expected - aperture) > 0.5:
            issue("FNumber/ApertureValue", "cross_field_inconsistency", f"expected_APEX={expected:.3f}, recorded={aperture:.3f}")
    iso_old, iso_new = number("ISOSpeedRatings"), number("ISOSpeed")
    if iso_old is not None and iso_new is not None and iso_old != iso_new:
        issue("ISOSpeedRatings/ISOSpeed", "cross_field_inconsistency", f"ratings={iso_old}, new_ISO={iso_new}")
    date_values = [clean_text(result.get(name, {}).get("raw_value")) for name in ("DateTime", "DateTimeOriginal", "DateTimeDigitized")]
    if all(date_values) and len(set(date_values)) > 1:
        issue("DateTime fields", "datetime_disagreement", " | ".join(date_values))


def statistical_outliers(metadata: pd.DataFrame, parameter_results: dict[str, dict[str, Any]], issues: list[dict[str, Any]]) -> None:
    continuous = ["ExposureTime", "FNumber", "ISOSpeedRatings", "FocalLength", "ExposureBiasValue", "BrightnessValue", "ShutterSpeedValue"]
    records = []
    for _, row in metadata.iterrows():
        patient_id = str(row["ID"])
        model = clean_text(parameter_results[patient_id].get("Model", {}).get("raw_value"))
        for parameter in continuous:
            item = parameter_results[patient_id].get(parameter, {})
            value = item.get("numeric_value")
            if value is not None and item.get("status") in {"valid", "warning"}:
                records.append({"ID": patient_id, "patient_group_id": row["patient_group_id"], "Model": model, "parameter": parameter, "value": value})
    frame = pd.DataFrame(records)
    for (model, parameter), group in frame.groupby(["Model", "parameter"]):
        if len(group) < 20 or group["value"].nunique() < 5:
            continue
        median = group["value"].median()
        mad = np.median(np.abs(group["value"] - median))
        if mad <= 0:
            continue
        score = 0.6745 * (group["value"] - median) / mad
        for row in group.loc[np.abs(score) > 5].itertuples(index=False):
            issues.append({"ID": row.ID, "patient_group_id": row.patient_group_id, "parameter": parameter, "issue_type": "device_stratified_statistical_outlier", "severity": "review", "raw_value": row.value, "detail": f"model={model}, median={median:.6g}, MAD={mad:.6g}"})


def plot_coverage(coverage: pd.DataFrame, path: Path) -> None:
    view = coverage.sort_values("usable_rate")
    fig, ax = plt.subplots(figsize=(10, 13))
    ax.barh(view["parameter"], view["nonblank_rate"] * 100, label="nonblank", color="#7AA6C2")
    ax.barh(view["parameter"], view["usable_rate"] * 100, label="valid/plausible", color="#2C7FB8")
    ax.set(xlabel="Images (%)", xlim=(0, 105), title="EXIF checklist completeness and plausibility")
    ax.legend()
    ax.grid(axis="x", alpha=0.2)
    fig.tight_layout()
    fig.savefig(path, dpi=200)
    plt.close(fig)


def write_report(path: Path, coverage: pd.DataFrame, image_audit: pd.DataFrame, patient_audit: pd.DataFrame, issues: pd.DataFrame, checklist: pd.DataFrame, group_source: str) -> None:
    shooting = coverage[coverage["section"] == "拍摄参数"]
    zero_nonblank = coverage.loc[coverage["nonblank_n"] == 0, "parameter"].tolist()
    incomplete = coverage.loc[coverage["nonblank_n"] < len(image_audit), ["parameter", "nonblank_n", "nonblank_rate"]].sort_values("nonblank_rate")
    invalid = coverage.loc[(coverage["invalid_n"] > 0) | (coverage["warning_n"] > 0), ["parameter", "invalid_n", "warning_n"]]
    ready_images = int(image_audit["core_research_ready"].sum())
    ready_patients = int(patient_audit["any_image_core_ready"].sum())
    all_direct_complete = int((image_audit["missing_or_blank_count_all42"] == 0).sum())
    all_shooting_complete = int((image_audit["missing_or_blank_count_shooting23"] == 0).sum())
    lines = [
        "# EXIF拍摄参数完整性与合理性审计",
        "",
        "> 本报告检查元数据字段是否存在、是否可解析、是否满足基础EXIF合法范围，并单独标注设备内统计离群值。统计离群或跨字段不一致是复核线索，不等同于源数据错误。",
        "",
        "## 数据概况",
        "",
        f"- 图片记录：{len(image_audit)}张；唯一图像ID：{image_audit['ID'].nunique()}。",
        f"- 患者组：{len(patient_audit)}；分组来源：{group_source}。",
        f"- 清单字段：{len(checklist)}项，其中常规EXIF {len(checklist)-len(shooting)}项、拍摄参数 {len(shooting)}项。",
        f"- 严格42项全部非空的图片：{all_direct_complete}张。",
        f"- 严格23项拍摄参数全部非空的图片：{all_shooting_complete}张。",
        f"- 满足研究核心参数规则的图片：{ready_images}张（{ready_images/len(image_audit):.1%}）；至少有一张合格图片的患者组：{ready_patients}组（{ready_patients/len(patient_audit):.1%}）。",
        "",
        "## 关键结论",
        "",
        "- 两个主要设备组合为 HONOR/BVL-AN00 与 Xiaomi/M2006J10C；多项字段的缺失具有明显设备系统性，不能把缺失简单解释为随机漏采。",
        "- `ISOSpeedRatings` 在全部图片中存在，可作为统一ISO主字段；`ISOSpeed` 是新版补充字段，不宜要求每张图片同时存在两者。",
        "- `FNumber`、`ExposureTime`、`FocalLength` 等核心直接拍摄参数覆盖完整；`ApertureValue`、`SensingMethod`、`DigitalZoomRatio` 等设备依赖字段不适合作为全队列必需条件。",
        "- `ImageDescription` 基本为空，虽然标签行存在，但没有可用于研究的值。",
        "- 正常范围判断同时输出明确非法值与待复核警告；后者包括设备编码、非信息性枚举码、跨字段不一致和设备内统计离群。",
    ]
    if zero_nonblank:
        lines += ["", f"- 完全没有非空值的字段：{', '.join(zero_nonblank)}。"]
    lines += [
        "",
        "## 全部存在且基础取值合法的拍摄参数",
        "",
        "以下14个拍摄参数在全部522张图片中均有非空记录，且单字段格式和基础EXIF取值范围检查全部通过。这里的“基础取值合法”不等于字段一定适合直接建模；设备依赖、变量缺乏变异或跨字段不一致仍需按说明处理。",
        "",
        "| 字段 | 中文说明 | 本数据取值概况 | 后续研究建议 |",
        "|---|---|---|---|",
        "| `ExposureTime` | 曝光时间，单位为秒，表示快门开启时长 | 0.002222–0.050009 s，中位数0.020 s | 推荐作为快门相关的主变量；比APEX形式更直观 |",
        "| `FNumber` | 光圈F值，反映镜头进光量和景深 | 1.89–2.00，中位数1.90 | 可直接使用，但其取值高度依赖设备型号 |",
        "| `ISOSpeedRatings` | ISO感光度，反映传感器增益 | 50–1600，中位数241.5 | 推荐作为统一ISO主字段；14张设备内高ISO离群图片应保留并复核 |",
        "| `FocalLength` | 实际焦距，通常以毫米为单位 | 1.82–6.67 mm，中位数6.67 mm | 可用于表征视角，但必须结合设备型号分析 |",
        "| `ExposureMode` | 曝光控制方式的EXIF枚举值 | 全部记录且枚举合法，本数据均为0（自动曝光） | 字段没有组内变异，不能单独提供预测信息 |",
        "| `MeteringMode` | 相机测量场景亮度的方式 | 合法值2或3 | 可作为分类变量，使用前应按EXIF标准解码 |",
        "| `Flash` | 闪光灯是否触发及工作模式的位掩码 | 合法值16或24 | 不能按连续数值处理，应解码或作为分类变量 |",
        "| `WhiteBalance` | 白平衡控制方式 | 全部为0（自动白平衡） | 无组内变异，不适合直接作为预测变量 |",
        "| `BrightnessValue` | 设备估计的场景亮度APEX值 | −2.67至10.90，中位数3.055 | 可用于环境亮度分析，但应做设备分层或标准化 |",
        "| `ShutterSpeedValue` | 快门速度的APEX表达 | 0–8.815，中位数5.058 | 字段本身合法，但239张Xiaomi图片与`ExposureTime`换算不一致；建模优先使用`ExposureTime` |",
        "| `SceneCaptureType` | 场景拍摄类型的EXIF枚举值 | 全部为0（标准场景） | 无组内变异，不适合直接作为预测变量 |",
        "| `SubsecTime` | `DateTime`对应的亚秒部分 | 全部为1–9位数字文本 | 主要用于时间精确匹配，不应作为连续拍摄参数直接建模 |",
        "| `SubsecTimeOriginal` | 原始拍摄时间的亚秒部分 | 522张均有合法数字文本 | 与`DateTimeOriginal`组合构成更精确时间戳 |",
        "| `SubsecTimeDigitized` | 数字化时间的亚秒部分 | 522张均有合法数字文本 | 主要用于时间一致性检查，通常不作为影像表型变量 |",
        "",
        "其中更适合进入后续跨设备拍摄参数研究的连续或有序核心变量是 `ExposureTime`、`FNumber`、`ISOSpeedRatings`、`FocalLength` 和 `BrightnessValue`。`MeteringMode`、`Flash` 可在正确解码后作为分类变量；`ExposureMode`、`WhiteBalance`、`SceneCaptureType` 在当前数据中没有变异，不能用于解释个体差异。",
    ]
    lines += ["", "## 字段覆盖率最低项", "", "| Parameter | Nonblank n | Nonblank rate |", "|---|---:|---:|"]
    for row in incomplete.head(15).itertuples(index=False):
        lines.append(f"| {row.parameter} | {int(row.nonblank_n)} | {row.nonblank_rate:.1%} |")
    lines += ["", "## 存在明确非法值或警告的字段", "", "| Parameter | Invalid n | Warning n |", "|---|---:|---:|"]
    for row in invalid.itertuples(index=False):
        lines.append(f"| {row.parameter} | {int(row.invalid_n)} | {int(row.warning_n)} |")
    lines += [
        "", "## 后续研究建议", "",
        "1. 建议使用统一核心变量集：设备厂商/型号、原始拍摄时间、曝光时间、FNumber、ISOSpeedRatings、实际焦距、曝光补偿、曝光/测光/闪光/白平衡模式、亮度值和场景类型。",
        "2. `ISOSpeed` 与 `ISOSpeedRatings` 合并为一个ISO变量，以Ratings优先或在一致时互补；不要将新版ISO字段缺失视为病例不合格。",
        "3. 35mm等效焦距为0、最大光圈APEX为0等值应按设备特异的无信息哨兵处理，而不是当作真实0值建模。",
        "4. 对设备系统性缺失字段，若进入模型必须增加缺失指示变量，并在设备分层或敏感性分析中验证；不能直接均值填补后忽略设备来源。",
        "5. 正式建模前优先复核 `parameter_value_issues.csv` 中 severity=invalid 的记录，再评估 review级离群值是否为真实拍摄差异。",
        "", "## 输出说明", "",
        "- `parameter_coverage.csv`：42项字段的覆盖、合法和警告统计。",
        "- `device_parameter_coverage.csv`：按相机型号分层的系统性缺失和合法性统计。",
        "- `image_parameter_audit.csv`：522张图片逐例完整性及研究可用性。",
        "- `patient_level_audit.csv`：按项目patient_group_id汇总的患者级可用性。",
        "- `parameter_value_issues.csv`：明确非法、设备编码、跨字段冲突和统计离群明细。",
        "- `parameter_values_long.csv`：逐图片逐参数的原始值、解析值和判断结果。",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run(args: argparse.Namespace) -> None:
    output = args.output_dir.resolve()
    output.mkdir(parents=True, exist_ok=True)
    checklist = pd.read_csv(args.checklist_csv, dtype=str, encoding="utf-8-sig")
    parameters = checklist["parameter"].tolist()
    metadata, raw_values, conflicts = load_workbook_data(args.metadata_xlsx, set(parameters))
    image_ids = set(metadata["ID"].astype(str))
    group_mapping, group_source = load_group_mapping(args.group_manifest, image_ids)
    metadata["patient_group_id"] = metadata["ID"].map(group_mapping)

    results: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    long_rows: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    checklist_index = checklist.set_index("parameter")
    for _, image in metadata.iterrows():
        patient_id = str(image["ID"])
        group_id = group_mapping[patient_id]
        for parameter in parameters:
            tag_present = parameter in raw_values.get(patient_id, {})
            raw_value = raw_values.get(patient_id, {}).get(parameter)
            if not tag_present:
                status, reason, numeric = "missing", "tag_not_present", None
            else:
                status, reason, numeric = validate_parameter(parameter, raw_value)
            item = {"raw_value": clean_text(raw_value), "tag_present": tag_present, "status": status, "reason": reason, "numeric_value": numeric}
            results[patient_id][parameter] = item
            row = checklist_index.loc[parameter]
            long_rows.append({"ID": patient_id, "patient_group_id": group_id, "section": row["section"], "parameter": parameter, "chinese_name": row["中文名称"], **item})
            if status in {"invalid", "warning"}:
                issues.append({"ID": patient_id, "patient_group_id": group_id, "parameter": parameter, "issue_type": reason, "severity": status, "raw_value": clean_text(raw_value), "detail": "field_rule"})
    values_long = pd.DataFrame(long_rows)
    model_by_id = {
        patient_id: results[patient_id]["Model"]["raw_value"]
        for patient_id in image_ids
    }
    values_long["Model"] = values_long["ID"].map(model_by_id)

    for _, row in metadata.iterrows():
        add_cross_field_issues(row, results, issues)
    statistical_outliers(metadata, results, issues)
    for (patient_id, parameter), values in conflicts.items():
        issues.append({"ID": patient_id, "patient_group_id": group_mapping.get(patient_id, patient_id), "parameter": parameter, "issue_type": "conflicting_duplicate_tag_values", "severity": "invalid", "raw_value": " | ".join(values), "detail": "multiple distinct values for same tag"})
    issues_frame = pd.DataFrame(issues).drop_duplicates()

    coverage_rows = []
    for parameter in parameters:
        subset = values_long[values_long["parameter"] == parameter]
        counts = subset["status"].value_counts()
        nonblank = int(subset["status"].isin(["valid", "warning", "invalid"]).sum())
        usable = int(subset["status"].isin(["valid", "warning"]).sum())
        numeric = pd.to_numeric(subset["numeric_value"], errors="coerce").dropna()
        coverage_rows.append({
            "section": checklist_index.loc[parameter, "section"], "parameter": parameter,
            "chinese_name": checklist_index.loc[parameter, "中文名称"], "total_images": len(metadata),
            "tag_present_n": int(subset["tag_present"].sum()), "tag_present_rate": float(subset["tag_present"].mean()),
            "nonblank_n": nonblank, "nonblank_rate": nonblank / len(metadata), "valid_n": int(counts.get("valid", 0)),
            "warning_n": int(counts.get("warning", 0)), "invalid_n": int(counts.get("invalid", 0)),
            "blank_n": int(counts.get("blank", 0)), "missing_tag_n": int(counts.get("missing", 0)),
            "usable_n": usable, "usable_rate": usable / len(metadata), "unique_nonblank_values": subset.loc[subset["status"].isin(["valid", "warning", "invalid"]), "raw_value"].nunique(),
            "numeric_min": numeric.min() if len(numeric) else math.nan, "numeric_median": numeric.median() if len(numeric) else math.nan,
            "numeric_max": numeric.max() if len(numeric) else math.nan,
            "most_common_values": json.dumps(subset.loc[subset["raw_value"] != "", "raw_value"].value_counts().head(8).to_dict(), ensure_ascii=False),
        })
    coverage = pd.DataFrame(coverage_rows)
    device_rows = []
    for (model, parameter), subset in values_long.groupby(["Model", "parameter"], dropna=False):
        counts = subset["status"].value_counts()
        device_rows.append({
            "Model": model,
            "parameter": parameter,
            "image_n": len(subset),
            "tag_present_n": int(subset["tag_present"].sum()),
            "nonblank_n": int(subset["status"].isin(["valid", "warning", "invalid"]).sum()),
            "valid_n": int(counts.get("valid", 0)),
            "warning_n": int(counts.get("warning", 0)),
            "invalid_n": int(counts.get("invalid", 0)),
            "missing_or_blank_n": int(counts.get("missing", 0) + counts.get("blank", 0)),
        })
    device_coverage = pd.DataFrame(device_rows)

    image_rows = []
    shooting_parameters = set(checklist.loc[checklist["section"] == "拍摄参数", "parameter"])
    issue_by_id = issues_frame.groupby("ID") if not issues_frame.empty else None
    for _, image in metadata.iterrows():
        patient_id = str(image["ID"])
        parameter_items = results[patient_id]
        missing_all = [name for name, item in parameter_items.items() if item["status"] in {"missing", "blank"}]
        invalid_all = [name for name, item in parameter_items.items() if item["status"] == "invalid"]
        warnings_all = [name for name, item in parameter_items.items() if item["status"] == "warning"]
        core_bad = [name for name in CORE_DIRECT if parameter_items[name]["status"] not in {"valid", "warning"}]
        iso_ready = any(parameter_items[name]["status"] in {"valid", "warning"} for name in ("ISOSpeedRatings", "ISOSpeed"))
        core_ready = not core_bad and iso_ready
        issue_subset = issue_by_id.get_group(patient_id) if issue_by_id is not None and patient_id in issue_by_id.groups else pd.DataFrame()
        image_rows.append({"ID": patient_id, "patient_group_id": group_mapping[patient_id], "filename": image["文件名"], "absolute_path": image["绝对路径"], "Make": parameter_items["Make"]["raw_value"], "Model": parameter_items["Model"]["raw_value"], "EXIF_present": image["EXIF存在"], "missing_or_blank_count_all42": len(missing_all), "missing_or_blank_fields_all42": ";".join(missing_all), "missing_or_blank_count_shooting23": len([name for name in missing_all if name in shooting_parameters]), "invalid_field_count": len(invalid_all), "invalid_fields": ";".join(invalid_all), "field_warning_count": len(warnings_all), "field_warning_fields": ";".join(warnings_all), "review_issue_count": int((issue_subset.get("severity", pd.Series(dtype=str)) == "review").sum()) if not issue_subset.empty else 0, "core_bad_fields": ";".join(core_bad), "ISO_fallback_ready": iso_ready, "core_research_ready": core_ready})
    image_audit = pd.DataFrame(image_rows)

    patient_rows = []
    for group_id, group in image_audit.groupby("patient_group_id"):
        patient_rows.append({"patient_group_id": group_id, "image_n": len(group), "image_ids": ";".join(group["ID"]), "device_make_n": group["Make"].nunique(), "device_model_n": group["Model"].nunique(), "device_models": ";".join(sorted(group["Model"].dropna().unique())), "any_image_core_ready": bool(group["core_research_ready"].any()), "all_images_core_ready": bool(group["core_research_ready"].all()), "core_ready_image_n": int(group["core_research_ready"].sum()), "total_invalid_fields": int(group["invalid_field_count"].sum()), "total_review_issues": int(group["review_issue_count"].sum()), "multi_image_device_conflict": bool(len(group) > 1 and group["Model"].nunique() > 1)})
    patient_audit = pd.DataFrame(patient_rows)

    save_csv(coverage, output / "parameter_coverage.csv")
    save_csv(device_coverage, output / "device_parameter_coverage.csv")
    save_csv(values_long, output / "parameter_values_long.csv")
    save_csv(image_audit, output / "image_parameter_audit.csv")
    save_csv(patient_audit, output / "patient_level_audit.csv")
    save_csv(issues_frame, output / "parameter_value_issues.csv")
    plot_coverage(coverage, output / "parameter_completeness.png")
    write_report(output / "exif_parameter_audit_report.md", coverage, image_audit, patient_audit, issues_frame, checklist, group_source)
    manifest = {"metadata_xlsx": str(args.metadata_xlsx.resolve()), "checklist_csv": str(args.checklist_csv.resolve()), "group_manifest": str(args.group_manifest.resolve()), "output_dir": str(output), "image_n": len(image_audit), "patient_group_n": len(patient_audit), "checklist_parameter_n": len(parameters), "core_ready_image_n": int(image_audit["core_research_ready"].sum()), "core_ready_patient_group_n": int(patient_audit["any_image_core_ready"].sum()), "issue_n": len(issues_frame), "created_at": datetime.now().isoformat()}
    (output / "run_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    print(coverage[["parameter", "nonblank_n", "valid_n", "warning_n", "invalid_n", "usable_rate"]].to_string(index=False))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--metadata-xlsx", type=Path, default=PROJECT_ROOT / "data" / "raw" / "EXIF" / "Image_Metadata_All.xlsx")
    parser.add_argument("--checklist-csv", type=Path, default=PROJECT_ROOT / "data" / "raw" / "EXIF" / "EXIF_Inform.csv")
    parser.add_argument("--group-manifest", type=Path, default=PROJECT_ROOT / "data" / "processed" / "splits" / "nyha_3class_sex_stratified_group_5fold.csv")
    parser.add_argument("--output-dir", type=Path, default=PROJECT_ROOT / "reports" / "exif_parameter_audit")
    return parser.parse_args()


if __name__ == "__main__":
    run(parse_args())
